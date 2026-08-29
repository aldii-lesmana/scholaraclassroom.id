# CARA MENJALANKAN:
#   1. pip install Flask Flask-MySQLdb
#   2. Import database.sql ke MySQL terlebih dahulu
#   3. Sesuaikan MYSQL_USER dan MYSQL_PASSWORD di bagian CONFIG
#   4. Jalankan: python app_sistem_penjualan.py
#   5. Buka browser: http://localhost:5000
#
# =============================================================================
# PANDUAN KOMENTAR BAHASA DALAM FILE INI:
#   [PYTHON]      → Kode backend Python / Flask
#   [HTML]        → Struktur markup halaman
#   [CSS]         → Gaya tampilan / styling
#   [JAVASCRIPT]  → Logika interaksi di browser
#   [SQL]         → Query / schema database MySQL
# =============================================================================


# =============================================================================
# [PYTHON] ── BAGIAN 1: IMPORT LIBRARY
# =============================================================================

from flask import Flask, render_template_string, render_template, request, redirect, url_for, flash, jsonify, send_file, session
# Flask               → framework web Python utama
# render_template_string → render HTML langsung dari string Python (tanpa file .html)
# request             → mengakses data dari form / URL
# redirect            → mengarahkan user ke halaman lain
# url_for             → membuat URL dari nama fungsi route
# flash               → mengirim pesan notifikasi sekali tampil
# jsonify             → mengubah dict Python menjadi respons JSON

import pymysql
import pymysql.cursors
from werkzeug.security import generate_password_hash, check_password_hash
# generate_password_hash -> enkripsi password aman (bcrypt/pbkdf2)
# check_password_hash    -> verifikasi password saat login
from functools import wraps
# wraps -> untuk membuat decorator login_required
from authlib.integrations.flask_client import OAuth
# OAuth  → library untuk koneksi Google OAuth 2.0
import secrets
import os
# secrets → untuk generate nonce yang aman
# MySQL               → ekstensi Flask untuk koneksi ke database MySQL

from datetime import date, datetime
# date                → tipe data tanggal (tahun-bulan-hari)
# datetime            → tipe data tanggal + waktu

import json
# json                → untuk encode/decode data JSON (dipakai di grafik Chart.js)

import io
# io                  -> membuat file di memori tanpa menyimpan ke disk (untuk download)

import smtplib
# smtplib             -> library bawaan Python untuk mengirim email via SMTP

from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
# email.mime          -> membuat struktur email (subject, body, attachment)

# [PYTHON] Library PDF -- install: pip install reportlab
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
# reportlab           -> library untuk membuat file PDF dari Python

# [PYTHON] Library Excel -- install: pip install openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
# openpyxl            -> library untuk membuat dan memformat file Excel (.xlsx)


# =============================================================================
# [PYTHON] ── BAGIAN 2: INISIALISASI APLIKASI FLASK
# =============================================================================

app = Flask(__name__)
app.config['DEBUG'] = True

# =============================================================================
# [PYTHON] ── BAGIAN 3: KONFIGURASI DATABASE & APLIKASI
# =============================================================================

# -- Kunci rahasia untuk session dan flash message --
app.config['SECRET_KEY'] = 'scholara_admin_secret_2024_very_secure'
app.config['SESSION_COOKIE_NAME'] = 'admin_session'
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_HTTPONLY'] = True

# ============================================================
# GOOGLE OAUTH CONFIG (Admin Panel - Port 5000)
# ============================================================
GOOGLE_CLIENT_ID     = os.environ.get('GOOGLE_CLIENT_ID', '')
GOOGLE_CLIENT_SECRET = os.environ.get('GOOGLE_CLIENT_SECRET', '')
ADMIN_GOOGLE_REDIRECT = os.environ.get('RENDER_EXTERNAL_URL', 'http://localhost:5000') + '/login/google/callback'

oauth = OAuth(app)
google_admin = oauth.register(
    name='google_admin',
    client_id=GOOGLE_CLIENT_ID,
    client_secret=GOOGLE_CLIENT_SECRET,
    server_metadata_url='https://accounts.google.com/.well-known/openid-configuration',
    client_kwargs={
        'scope': 'openid email profile',
    }
)

# -- Konfigurasi koneksi MySQL --
import os
import uuid

# -- Konfigurasi upload gambar produk --
BASE_DIR      = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'static', 'uploads', 'produk')
ALLOWED_EXT   = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def allowed_file(filename):
    """[PYTHON] Cek apakah ekstensi file termasuk yang diizinkan."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXT

def save_gambar(file):
    """
    [PYTHON] Simpan file gambar ke folder uploads.
    Cek nama file asli (spasi) DAN nama secure (underscore).
    Jika salah satunya sudah ada di folder, langsung pakai — TIDAK buat duplikat.
    """
    from werkzeug.utils import secure_filename
    if file and allowed_file(file.filename):
        os.makedirs(UPLOAD_FOLDER, exist_ok=True)

        original_name = file.filename                   # contoh: "mouse bluethoot.png"
        secure_name   = secure_filename(file.filename)  # contoh: "mouse_bluethoot.png"

        # Cek nama asli dulu (file yang sudah manual ditaruh di folder)
        if os.path.exists(os.path.join(UPLOAD_FOLDER, original_name)):
            return original_name   # pakai file asli, tidak save ulang

        # Cek nama secure (sudah pernah diupload via panel)
        if os.path.exists(os.path.join(UPLOAD_FOLDER, secure_name)):
            return secure_name     # pakai file yang ada, tidak save ulang

        # Belum ada sama sekali — baru simpan dengan nama secure
        file.save(os.path.join(UPLOAD_FOLDER, secure_name))
        return secure_name
    return None

app.config['MYSQL_HOST']     = '127.0.0.1'
app.config['MYSQL_USER']     = 'root'
app.config['MYSQL_PASSWORD'] = ''
app.config['MYSQL_DB']       = 'sistem_penjualan'
app.config['MYSQL_PORT']     = 3306
app.config['MYSQL_CURSORCLASS'] = 'DictCursor'

# PyMySQL helper — pengganti Flask-MySQLdb yang lebih stabil
class MySQL:
    def __init__(self, app=None):
        self.app = app

    def get_connection(self):
        return pymysql.connect(
            host=app.config['MYSQL_HOST'],
            user=app.config['MYSQL_USER'],
            password=app.config['MYSQL_PASSWORD'],
            database=app.config['MYSQL_DB'],
            port=app.config['MYSQL_PORT'],
            cursorclass=pymysql.cursors.DictCursor,
            autocommit=False,
            charset='utf8mb4'
        )

    @property
    def connection(self):
        from flask import g
        if 'mysql_conn' not in g or not g.mysql_conn.open:
            g.mysql_conn = self.get_connection()
        return g.mysql_conn

mysql = MySQL(app)

@app.teardown_appcontext
def close_mysql_connection(exception=None):
    from flask import g
    conn = g.pop('mysql_conn', None)
    if conn is not None:
        conn.close()


# =============================================================================
# [SQL] ── BAGIAN 4: SCHEMA DATABASE (jalankan ini di MySQL terlebih dahulu)
# =============================================================================
# 
# Salin kode SQL di bawah ini dan jalankan di MySQL Workbench / phpMyAdmin:
#
# CREATE DATABASE IF NOT EXISTS sistem_penjualan;
# USE sistem_penjualan;
#
# -- Tabel pelanggan: menyimpan data pelanggan / klien --
# CREATE TABLE pelanggan (
#     id              INT AUTO_INCREMENT PRIMARY KEY,
#     kode_pelanggan  VARCHAR(20) UNIQUE NOT NULL,     -- kode unik: PLG-001
#     nama            VARCHAR(100) NOT NULL,            -- nama lengkap / perusahaan
#     email           VARCHAR(100),                     -- email kontak
#     telepon         VARCHAR(20),                      -- nomor telepon
#     alamat          TEXT,                             -- alamat lengkap
#     kota            VARCHAR(50),                      -- kota domisili
#     status          ENUM('aktif','tidak_aktif') DEFAULT 'aktif',
#     created_at      TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
#     updated_at      TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
# );
#
# -- Tabel kategori: pengelompokan produk --
# CREATE TABLE kategori (
#     id          INT AUTO_INCREMENT PRIMARY KEY,
#     nama        VARCHAR(50) NOT NULL,
#     deskripsi   TEXT
# );
#
# -- Tabel produk: data produk beserta stok --
# CREATE TABLE produk (
#     id              INT AUTO_INCREMENT PRIMARY KEY,
#     kode_produk     VARCHAR(20) UNIQUE NOT NULL,     -- kode unik: PRD-001
#     nama            VARCHAR(100) NOT NULL,
#     kategori_id     INT,
#     harga           DECIMAL(15,2) NOT NULL,           -- harga jual satuan
#     stok            INT DEFAULT 0,                    -- jumlah stok tersedia
#     stok_minimum    INT DEFAULT 5,                    -- batas peringatan stok menipis
#     satuan          VARCHAR(20) DEFAULT 'pcs',        -- satuan: pcs, unit, kg, dll
#     deskripsi       TEXT,
#     status          ENUM('aktif','tidak_aktif') DEFAULT 'aktif',
#     created_at      TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
#     updated_at      TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
#     FOREIGN KEY (kategori_id) REFERENCES kategori(id)
# );
#
# -- Tabel pesanan: header transaksi penjualan --
# CREATE TABLE pesanan (
#     id              INT AUTO_INCREMENT PRIMARY KEY,
#     no_pesanan      VARCHAR(20) UNIQUE NOT NULL,     -- nomor unik: ORD-001
#     pelanggan_id    INT NOT NULL,
#     tanggal_pesan   DATE NOT NULL,
#     tanggal_kirim   DATE,
#     status          ENUM('pending','diproses','dikirim','selesai','dibatalkan') DEFAULT 'pending',
#     total_harga     DECIMAL(15,2) DEFAULT 0,          -- total sebelum diskon
#     diskon          DECIMAL(5,2) DEFAULT 0,           -- persentase diskon (0-100)
#     total_bayar     DECIMAL(15,2) DEFAULT 0,          -- total setelah diskon
#     catatan         TEXT,
#     created_at      TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
#     updated_at      TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
#     FOREIGN KEY (pelanggan_id) REFERENCES pelanggan(id)
# );
#
# -- Tabel detail_pesanan: baris item dalam satu pesanan --
# CREATE TABLE detail_pesanan (
#     id              INT AUTO_INCREMENT PRIMARY KEY,
#     pesanan_id      INT NOT NULL,
#     produk_id       INT NOT NULL,
#     jumlah          INT NOT NULL,
#     harga_satuan    DECIMAL(15,2) NOT NULL,
#     subtotal        DECIMAL(15,2) NOT NULL,           -- harga_satuan × jumlah
#     FOREIGN KEY (pesanan_id) REFERENCES pesanan(id) ON DELETE CASCADE,
#     FOREIGN KEY (produk_id) REFERENCES produk(id)
# );
#
# -- Data awal: kategori --
# INSERT INTO kategori (nama) VALUES ('Elektronik'),('Pakaian'),('Makanan'),('Alat Tulis');
#
# -- Data awal: pelanggan contoh --
# INSERT INTO pelanggan (kode_pelanggan,nama,email,telepon,kota,status) VALUES
# ('PLG-001','PT Maju Jaya','maju@email.com','021-111','Jakarta','aktif'),
# ('PLG-002','CV Berkah','berkah@email.com','022-222','Bandung','aktif');
#
# -- Data awal: produk contoh --
# INSERT INTO produk (kode_produk,nama,kategori_id,harga,stok,stok_minimum,satuan) VALUES
# ('PRD-001','Laptop Asus',1,8500000,15,5,'unit'),
# ('PRD-002','Mouse Wireless',1,250000,50,10,'pcs'),
# ('PRD-003','Kaos Polos',2,85000,100,20,'pcs'),
# ('PRD-004','Pulpen Pilot',4,8000,3,30,'pcs');
# =============================================================================


# =============================================================================
# [PYTHON] ── BAGIAN 5: FUNGSI PEMBANTU (HELPER FUNCTIONS)
# =============================================================================

def generate_kode(prefix, table, kolom):
    """
    [PYTHON] Menghasilkan kode unik otomatis, misal: PLG-001, PRD-002, ORD-003
    
    Parameter:
        prefix  → awalan kode (PLG, PRD, ORD)
        table   → nama tabel di database
        kolom   → nama kolom kode di tabel tersebut
    """
    # [PYTHON] Cari nomor urutan terakhir yang valid (format PREFIX-NNN)
    # Ambil semua kode lalu cari yang bisa di-parse sebagai angka
    cur = mysql.connection.cursor()
    cur.execute(f"SELECT {kolom} FROM {table} ORDER BY id DESC")
    rows = cur.fetchall()
    cur.close()

    last_num = 0
    for row in rows:
        try:
            parts = row[kolom].split('-')
            if len(parts) >= 2 and parts[0] == prefix:
                num = int(parts[1])   # hanya parse jika prefix cocok dan suffix angka
                if num > last_num:
                    last_num = num
        except (ValueError, AttributeError):
            continue  # abaikan baris dengan format kode yang tidak valid

    # Jika tidak ada kode valid ditemukan, hitung jumlah baris sebagai fallback
    if last_num == 0:
        cur2 = mysql.connection.cursor()
        cur2.execute(f"SELECT COUNT(*) as total FROM {table}")
        result = cur2.fetchone()
        cur2.close()
        last_num = (result['total'] if result else 0)

    return f"{prefix}-{str(last_num + 1).zfill(3)}"   # zfill(3) → padding nol: 1 → "001"


def format_rupiah(value):
    """
    [PYTHON] Filter Jinja2: mengubah angka menjadi format mata uang Rupiah
    Contoh: 8500000 → 'Rp 8.500.000'
    """
    return f"Rp {value:,.0f}".replace(',', '.')                 # :,.0f → format ribuan dengan koma


# Mendaftarkan fungsi format_rupiah sebagai filter di template Jinja2
app.jinja_env.filters['rupiah'] = format_rupiah

# Mendaftarkan fungsi enumerate bawaan Python agar bisa dipakai di template Jinja2
app.jinja_env.filters['enumerate'] = enumerate


@app.context_processor
def inject_now():
    """
    [PYTHON] Context processor: menyuntikkan variabel 'now' dan 'format_rupiah' ke semua template
    Sehingga {{ now.strftime('%d %b %Y') }} dan {{ format_rupiah(harga) }} bisa dipakai di semua halaman HTML
    """
    return {
        'now': datetime.now(),
        'format_rupiah': format_rupiah,
    }


def login_required(f):
    """
    [PYTHON] Sistem login dinonaktifkan — semua route bisa diakses langsung.
    Decorator ini dibiarkan ada agar kode tidak error, tapi tidak memblokir akses.
    """
    @wraps(f)
    def decorated(*args, **kwargs):
        return f(*args, **kwargs)   # langsung teruskan tanpa cek session
    return decorated


def get_current_user():
    """[PYTHON] Sistem login dinonaktifkan — selalu return None."""
    return None


# =============================================================================
# [PYTHON] -- HALAMAN LOGIN / LOGOUT / REGISTER
# Route ini ada agar tidak 404. Karena sistem login dinonaktifkan,
# semua redirect langsung ke /dashboard.
# =============================================================================

@app.route('/login', methods=['GET', 'POST'])
def login():
    """[PYTHON] Sistem login dinonaktifkan — langsung redirect ke dashboard."""
    return redirect('/dashboard')

@app.route('/register', methods=['GET', 'POST'])
def register():
    """[PYTHON] Sistem login dinonaktifkan — langsung redirect ke dashboard."""
    return redirect('/dashboard')

@app.route('/logout')
def logout():
    """[PYTHON] Sistem login dinonaktifkan — redirect ke dashboard."""
    session.clear()
    return redirect('/dashboard')

@app.route('/login/google')
def login_google():
    """[PYTHON] Sistem login Google dinonaktifkan — redirect ke dashboard."""
    return redirect('/dashboard')

@app.route('/login/google/callback')
def login_google_callback():
    """[PYTHON] Sistem login Google dinonaktifkan — redirect ke dashboard."""
    return redirect('/dashboard')


# =============================================================================
# [PYTHON] ── BAGIAN 6: TEMPLATE HTML/CSS/JS (Di-embed sebagai string Python)
# =============================================================================
# Semua tampilan antarmuka ada di sini.
# Format: render_template_string(NAMA_TEMPLATE, variabel=nilai)
# =============================================================================


# -----------------------------------------------------------------------------
# [CSS] Template CSS global — dipakai di semua halaman
# -----------------------------------------------------------------------------

# -----------------------------------------------------------------------------
# [PYTHON + HTML] Template dasar: sidebar + topbar yang dipakai semua halaman
# Menggunakan Jinja2 template syntax: {{ variabel }}, {% blok %}
# -----------------------------------------------------------------------------

# -----------------------------------------------------------------------------
# [PYTHON] Fungsi pembantu render template: menyatukan base + konten halaman
# -----------------------------------------------------------------------------
def render_page(page_title, active_menu, content, extra_scripts=""):
    """
    [PYTHON] Menggabungkan BASE_TEMPLATE dengan konten halaman tertentu.
    
    Parameter:
        page_title    → judul halaman (ditampilkan di tab dan topbar)
        active_menu   → nama menu aktif untuk highlight sidebar
        content       → konten HTML halaman (dari template masing-masing)
        extra_scripts → JavaScript tambahan (misal: inisialisasi Chart.js)
    """
    return render_template(
        "admin/base.html",            # template dasar dengan sidebar dan topbar
        page_title=page_title,    # judul halaman
        active_menu=active_menu,  # menu yang sedang aktif
        content=content,          # konten halaman
        extra_scripts=extra_scripts, # script tambahan
        now=datetime.now()        # waktu sekarang untuk topbar
    )


# =============================================================================
# [PYTHON] ── BAGIAN 7: ROUTE DASHBOARD
# Route adalah URL yang dikaitkan dengan fungsi Python menggunakan @app.route
# =============================================================================

@app.route('/')              # URL root '/' → redirect ke /dashboard
@login_required
def index():
    return redirect('/dashboard')  # redirect ke halaman dashboard


@app.route('/dashboard')     # URL /dashboard → tampilkan halaman dashboard
@login_required
def dashboard():
    cur = mysql.connection.cursor()
    cur.execute("SELECT COUNT(*) as total FROM pelanggan WHERE status='aktif'")
    total_pelanggan = cur.fetchone()['total']
    cur.execute("SELECT COUNT(*) as total FROM produk WHERE status='aktif'")
    total_produk = cur.fetchone()['total']
    cur.execute("SELECT COUNT(*) as total FROM pesanan WHERE MONTH(tanggal_pesan) = MONTH(NOW()) AND YEAR(tanggal_pesan) = YEAR(NOW())")
    total_pesanan = cur.fetchone()['total']
    cur.execute("SELECT COALESCE(SUM(total_bayar), 0) as total FROM pesanan WHERE status = 'selesai' AND MONTH(tanggal_pesan) = MONTH(NOW()) AND YEAR(tanggal_pesan) = YEAR(NOW())")
    total_penjualan = cur.fetchone()['total']
    cur.execute("SELECT p.no_pesanan, pl.nama as nama_pelanggan, p.tanggal_pesan, p.total_bayar, p.status FROM pesanan p JOIN pelanggan pl ON p.pelanggan_id = pl.id ORDER BY p.created_at DESC LIMIT 5")
    pesanan_terbaru = cur.fetchall()
    cur.execute("SELECT nama, stok, stok_minimum, satuan FROM produk WHERE stok <= stok_minimum AND status = 'aktif'")
    stok_menipis = cur.fetchall()
    cur.execute("SELECT DATE_FORMAT(tanggal_pesan, '%b %Y') as bulan, COALESCE(SUM(total_bayar), 0) as total FROM pesanan WHERE status = 'selesai' AND tanggal_pesan >= DATE_SUB(NOW(), INTERVAL 6 MONTH) GROUP BY DATE_FORMAT(tanggal_pesan, '%Y-%m') ORDER BY tanggal_pesan ASC")
    grafik_data = cur.fetchall()
    cur.execute("SELECT status, COUNT(*) as jumlah FROM pesanan WHERE MONTH(tanggal_pesan) = MONTH(NOW()) GROUP BY status")
    status_data = cur.fetchall()
    cur.close()
    
    import json
    grafik_json = json.dumps(list(grafik_data), default=str)
    status_json = json.dumps(list(status_data), default=str)

    return render_template('admin/dashboard.html', page_title='Dashboard', active_menu='dashboard',
                           total_pelanggan=total_pelanggan, total_produk=total_produk,
                           total_pesanan=total_pesanan, total_penjualan=total_penjualan,
                           pesanan_terbaru=pesanan_terbaru, stok_menipis=stok_menipis,
                           grafik_json=grafik_json, status_json=status_json)


# [PYTHON] Fungsi pembantu untuk merender baris tabel pesanan terbaru



# [PYTHON] Fungsi pembantu untuk merender tabel stok menipis
def _render_stok_menipis(data):
    """[PYTHON] Menghasilkan HTML tabel atau pesan kosong untuk stok menipis."""
    if not data:
        return '<div class="card-body"><div class="empty-state"><div class="empty-icon">✓</div><p>Stok semua aman</p></div></div>'
    
    rows = ''
    for s in data:
        rows += f"""<tr>
            <td>{s['nama']}</td>
            <td><span class="badge badge-red">{s['stok']} {s['satuan']}</span></td>
            <td class="td-mono">{s['stok_minimum']}</td>
        </tr>"""
    return f"""<table>
        <thead><tr><th>Produk</th><th>Stok</th><th>Minimum</th></tr></thead>
        <tbody>{rows}</tbody>
    </table>"""


# =============================================================================
# [PYTHON] ── BAGIAN 8: ROUTE PELANGGAN (CRUD)
# CRUD = Create (tambah), Read (tampilkan), Update (edit), Delete (hapus)
# =============================================================================

@app.route('/pelanggan')   # menampilkan daftar semua pelanggan
@login_required
def pelanggan_index():
    cur = mysql.connection.cursor()
    search = request.args.get('search', '')
    query = """
        SELECT pl.*,
            COUNT(DISTINCT p.id)              AS total_pesanan,
            COALESCE(SUM(dp.jumlah), 0)       AS total_item,
            COALESCE(SUM(p.total_bayar), 0)   AS total_belanja,
            GROUP_CONCAT(DISTINCT pr.nama SEPARATOR ', ') AS nama_produk_list
        FROM pelanggan pl
        LEFT JOIN pesanan p   ON pl.id = p.pelanggan_id
        LEFT JOIN detail_pesanan dp ON p.id = dp.pesanan_id
        LEFT JOIN produk pr   ON dp.produk_id = pr.id
        WHERE 1=1
    """
    params = []
    if search:
        query += " AND (pl.kode_pelanggan LIKE %s OR pl.nama LIKE %s)"
        params += [f'%{search}%', f'%{search}%']
    query += " GROUP BY pl.id ORDER BY pl.id DESC"
    cur.execute(query, params)
    data = cur.fetchall()
    cur.close()
    return render_template('admin/pelanggan_index.html', page_title='Pelanggan', active_menu='pelanggan', data=data, search=search)


@app.route('/pelanggan/tambah', methods=['GET', 'POST'])
@login_required
# methods=['GET','POST']: route ini menerima 2 metode HTTP
# GET  → menampilkan form kosong
# POST → memproses data yang dikirim dari form
def pelanggan_tambah():
    if request.method == 'POST':
        kode = generate_kode('PLG', 'pelanggan', 'kode_pelanggan')
        nama, email = request.form['nama'], request.form['email']
        telepon, alamat = request.form['telepon'], request.form['alamat']
        kota, status = request.form['kota'], request.form.get('status', 'aktif')
        cur = mysql.connection.cursor()
        cur.execute("INSERT INTO pelanggan (kode_pelanggan, nama, email, telepon, alamat, kota, status) VALUES (%s, %s, %s, %s, %s, %s, %s)", (kode, nama, email, telepon, alamat, kota, status))
        mysql.connection.commit()
        cur.close()
        flash(f"Pelanggan {nama} berhasil ditambahkan!", "success")
        return redirect('/pelanggan')
    return render_template('admin/pelanggan_form.html', page_title='Tambah Pelanggan', active_menu='pelanggan', is_edit=False, data=None)


@app.route('/pelanggan/edit/<int:id>', methods=['GET', 'POST'])
@login_required
# <int:id>: parameter URL dinamis — id dikonversi ke integer
def pelanggan_edit(id):
    """[PYTHON] Menampilkan form edit dan menyimpan perubahan data pelanggan."""
    cur = mysql.connection.cursor()

    if request.method == 'POST':
        # [PYTHON] Ambil semua field dari form
        nama    = request.form['nama']
        email   = request.form['email']
        telepon = request.form['telepon']
        alamat  = request.form['alamat']
        kota    = request.form['kota']
        status  = request.form['status']

        # [SQL] UPDATE: memperbarui data yang sudah ada berdasarkan id
        cur.execute("""
            UPDATE pelanggan
            SET nama=%s, email=%s, telepon=%s, alamat=%s, kota=%s, status=%s
            WHERE id=%s
        """, (nama, email, telepon, alamat, kota, status, id))
        mysql.connection.commit()
        cur.close()
        flash('Data pelanggan berhasil diperbarui!', 'success')
        return redirect('/pelanggan')

    # [SQL] SELECT: ambil data pelanggan berdasarkan id untuk mengisi form
    cur.execute("SELECT * FROM pelanggan WHERE id=%s", (id,))
    data = cur.fetchone()  # fetchone() karena hanya 1 baris yang dibutuhkan
    cur.close()

    if not data:
        flash('Data tidak ditemukan!', 'danger')
        return redirect('/pelanggan')

    # [PYTHON] Tentukan opsi 'selected' untuk dropdown status
    opt_aktif    = 'selected' if data['status'] == 'aktif' else ''
    opt_nonaktif = 'selected' if data['status'] == 'tidak_aktif' else ''

    # [HTML] Template form edit — nilai input diisi dengan data dari database
    content = f"""
    <div class="page-header">
        <div>
            <h2>Edit Pelanggan</h2>
            <div class="breadcrumb"><a href="/pelanggan">Pelanggan</a> / Edit — {data['kode_pelanggan']}</div>
        </div>
    </div>
    <div class="form-card" style="max-width:680px">
        <div class="form-section-title">Informasi Pelanggan</div>
        <form method="POST">
            <div class="form-grid">
                <div class="form-group full">
                    <label>Nama *</label>
                    <!-- [HTML] value: mengisi input dengan data yang sudah ada -->
                    <input type="text" name="nama" value="{data['nama']}" required>
                </div>
                <div class="form-group">
                    <label>Email</label>
                    <input type="email" name="email" value="{data['email'] or ''}">
                </div>
                <div class="form-group">
                    <label>Telepon</label>
                    <input type="text" name="telepon" value="{data['telepon'] or ''}">
                </div>
                <div class="form-group">
                    <label>Kota</label>
                    <input type="text" name="kota" value="{data['kota'] or ''}">
                </div>
                <div class="form-group">
                    <label>Status</label>
                    <!-- [HTML] select: dropdown pilihan dengan opsi -->
                    <select name="status">
                        <option value="aktif" {opt_aktif}>Aktif</option>
                        <option value="tidak_aktif" {opt_nonaktif}>Tidak Aktif</option>
                    </select>
                </div>
                <div class="form-group full">
                    <label>Alamat</label>
                    <textarea name="alamat">{data['alamat'] or ''}</textarea>
                </div>
            </div>
            <div class="form-actions">
                <button type="submit" class="btn btn-primary">Simpan Perubahan</button>
                <a href="/pelanggan" class="btn btn-outline">Batal</a>
            </div>
        </form>
    </div>
    """
    return render_page('Edit Pelanggan', 'pelanggan', content)


@app.route('/pelanggan/hapus/<int:id>')
@login_required
def pelanggan_hapus(id):
    """
    [PYTHON] Soft delete: tidak menghapus data, hanya mengubah status menjadi 'tidak_aktif'.
    Ini praktik terbaik agar histori data tetap terjaga.
    """
    cur = mysql.connection.cursor()
    # [SQL] UPDATE status menjadi tidak_aktif (soft delete)
    cur.execute("UPDATE pelanggan SET status='tidak_aktif' WHERE id=%s", (id,))
    mysql.connection.commit()
    cur.close()
    flash('Pelanggan berhasil dinonaktifkan!', 'warning')
    return redirect('/pelanggan')


# =============================================================================
# [PYTHON] ── BAGIAN 9: ROUTE PRODUK (CRUD)
# =============================================================================

@app.route('/produk')
@login_required
def produk_index():
    """[PYTHON] Menampilkan daftar produk beserta kategori dan status stok."""
    cur = mysql.connection.cursor()
    search = request.args.get('search', '')

    if search:
        # [SQL] LEFT JOIN: tampilkan semua produk, dan nama kategori jika ada
        # LEFT JOIN tetap menampilkan produk meski tidak punya kategori (kategori_id = NULL)
        cur.execute("""
            SELECT p.*, k.nama as nama_kategori
            FROM produk p LEFT JOIN kategori k ON p.kategori_id = k.id
            WHERE p.nama LIKE %s OR p.kode_produk LIKE %s
            ORDER BY p.id DESC
        """, (f'%{search}%', f'%{search}%'))
    else:
        cur.execute("""
            SELECT p.*, k.nama as nama_kategori
            FROM produk p LEFT JOIN kategori k ON p.kategori_id = k.id
            ORDER BY p.id DESC
        """)

    data = cur.fetchall()
    cur.close()

    return render_template('admin/produk_index.html', page_title='Produk & Stok', active_menu='produk', data=data, search=search)



@app.route('/produk/tambah', methods=['GET', 'POST'])
@login_required
def produk_tambah():
    """[PYTHON] Form tambah produk baru dengan upload foto dan pilihan kategori."""
    cur = mysql.connection.cursor()

    if request.method == 'POST':
        kode         = generate_kode('PRD', 'produk', 'kode_produk')
        nama         = request.form['nama']
        kategori_id  = request.form['kategori_id'] or None  # None jika tidak dipilih
        # Hapus titik dan koma dari input angka
        harga        = request.form['harga'].replace('.', '').replace(',', '')
        stok         = request.form['stok'].replace('.', '').replace(',', '') if request.form['stok'] else 0
        stok_minimum = request.form['stok_minimum'].replace('.', '').replace(',', '') if request.form['stok_minimum'] else 0
        satuan       = request.form['satuan']
        deskripsi    = request.form.get('deskripsi', '')  # .get() → tidak error jika kosong
        mapel        = request.form.get('mapel', '') or None
        penerbit     = request.form.get('penerbit', '') or None
        kurikulum    = request.form.get('kurikulum', '') or None
        kelas        = request.form.get('kelas', '') or None
        semester     = request.form.get('semester', '') or None

        # [PYTHON] Proses upload gambar jika ada
        gambar = None
        if 'gambar' in request.files:
            file = request.files['gambar']
            if file and file.filename:
                gambar = save_gambar(file)

        # [SQL] Tambahkan kolom secara otomatis jika belum ada di database
        for col in [
            "gambar VARCHAR(255) NULL",
            "mapel VARCHAR(100) NULL",
            "penerbit VARCHAR(100) NULL",
            "kurikulum VARCHAR(100) NULL",
            "kelas VARCHAR(50) NULL",
            "semester VARCHAR(50) NULL"
        ]:
            try:
                cur.execute(f"ALTER TABLE produk ADD COLUMN {col}")
                mysql.connection.commit()
            except:
                pass # Abaikan jika kolom sudah ada

        # [SQL] INSERT produk baru
        cur.execute("""
            INSERT INTO produk (kode_produk, nama, kategori_id, harga, stok, stok_minimum, satuan, deskripsi, gambar, mapel, penerbit, kurikulum, kelas, semester)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (kode, nama, kategori_id, harga, stok, stok_minimum, satuan, deskripsi, gambar, mapel, penerbit, kurikulum, kelas, semester))
        mysql.connection.commit()
        cur.close()
        flash('Produk berhasil ditambahkan!', 'success')
        return redirect('/produk')

    # [SQL] Ambil semua kategori untuk dropdown
    cur.execute("SELECT * FROM kategori ORDER BY id")
    kategori = cur.fetchall()
    cur.close()

    # [PYTHON] Bangun opsi dropdown kategori — pisahkan jenjang pendidikan
    JENJANG_NAMA = ('SD', 'SMP', 'SMA', 'SMK', 'Perguruan Tinggi')
    opt_kategori  = '<option value="">&#8212; Pilih Kategori &#8212;</option>'
    opt_kategori += '<optgroup label="&#127891; Jenjang Pendidikan">'
    for k in kategori:
        if k['nama'] in JENJANG_NAMA:
            opt_kategori += f'<option value="{k["id"]}">{k["nama"]}</option>'
    opt_kategori += '</optgroup>'
    opt_kategori += '<optgroup label="Kategori Lainnya">'
    for k in kategori:
        if k['nama'] not in JENJANG_NAMA:
            opt_kategori += f'<option value="{k["id"]}">{k["nama"]}</option>'
    opt_kategori += '</optgroup>'

    import json as _json
    kategori_jenjang_map = {str(k['id']): k['nama'] for k in kategori if k['nama'] in JENJANG_NAMA}

    content = f"""
    <div class="page-header">
        <div><h2>Tambah Produk</h2>
        <div class="breadcrumb"><a href="/produk">Produk</a> / Tambah</div></div>
    </div>
    <div class="form-card" style="max-width:720px">
        <div class="form-section-title">Informasi Produk</div>
        <form method="POST" enctype="multipart/form-data">
            <div class="form-grid">
                <div class="form-group full">
                    <label>Nama Produk *</label>
                    <input type="text" name="nama" required placeholder="Nama produk">
                </div>
                <div class="form-group">
                    <label>Kategori</label>
                    <select name="kategori_id" id="kategori_id_tambah" onchange="toggleBukuFields(this, 'tambah')">{opt_kategori}</select>
                </div>
                <div class="form-group">
                    <label>Satuan</label>
                    <select name="satuan">
                        <option value="pcs">pcs</option>
                        <option value="buku">buku</option>
                        <option value="unit">unit</option>
                        <option value="kg">kg</option>
                        <option value="liter">liter</option>
                        <option value="box">box</option>
                        <option value="lusin">lusin</option>
                    </select>
                </div>
                <div class="form-group">
                    <label>Harga Jual (Rp) *</label>
                    <input type="number" name="harga" required placeholder="0" min="0">
                </div>
                <div class="form-group">
                    <label>Stok Awal</label>
                    <input type="number" name="stok" value="0" min="0">
                </div>
                <div class="form-group">
                    <label>Stok Minimum (peringatan)</label>
                    <input type="number" name="stok_minimum" value="5" min="0">
                </div>
                <div class="form-group full">
                    <label>Deskripsi</label>
                    <textarea name="deskripsi" placeholder="Opsional \u2014 deskripsi singkat produk"></textarea>
                </div>
                <!-- [HTML] Drag-and-drop upload foto produk -->
                <div class="form-group full">
                    <label>Foto Produk</label>
                    <div id="drop-zone-tambah"
                         style="border:2px dashed var(--border,#e2e8f0);border-radius:10px;padding:32px 20px;text-align:center;
                                cursor:pointer;transition:border-color .2s,background .2s;background:var(--bg-2,#f8fafc);"
                         onclick="document.getElementById('gambar_tambah').click()"
                         ondragover="event.preventDefault();this.style.borderColor='var(--primary,#2d6a4f)';this.style.background='var(--bg-3,#f0fdf4)'"
                         ondragleave="this.style.borderColor='';this.style.background=''"
                         ondrop="handleDrop(event,'tambah')">
                        <div id="preview-tambah" style="display:none;">
                            <img id="preview-img-tambah" src="" style="max-height:160px;max-width:100%;border-radius:8px;object-fit:contain;">
                            <p id="preview-name-tambah" style="margin:8px 0 0;font-size:13px;color:var(--text-2,#64748b);"></p>
                        </div>
                        <div id="drop-placeholder-tambah">
                            <div style="font-size:36px;margin-bottom:8px;">&#128247;</div>
                            <p style="margin:0;font-weight:600;font-size:14px;">Drag &amp; drop foto ke sini</p>
                            <p style="margin:4px 0 12px;font-size:12px;color:var(--text-3,#94a3b8);">atau klik untuk memilih file</p>
                            <span style="font-size:11px;color:var(--text-3,#94a3b8);">PNG, JPG, JPEG, WEBP, GIF &mdash; maks. 5 MB</span>
                        </div>
                    </div>
                    <input type="file" name="gambar" id="gambar_tambah" accept="image/*" style="display:none"
                           onchange="previewFile(this,'tambah')">
                </div>
            </div>

            <!-- SECTION INFORMASI BUKU -->
            <div id="buku-fields-tambah" style="display:none;margin-top:24px;padding-top:20px;border-top:2px solid #e2f0ff;">
                <div style="display:flex;align-items:center;gap:8px;margin-bottom:16px;">
                    <span style="font-size:20px;">&#128218;</span>
                    <div class="form-section-title" style="margin:0;">Informasi Buku</div>
                    <span style="font-size:11px;color:var(--text-3);margin-left:4px;">(opsional)</span>
                </div>
                <div class="form-grid">
                    <div class="form-group">
                        <label>Mata Pelajaran</label>
                        <input type="text" name="mapel" id="mapel_tambah" placeholder="Misal: Matematika, Fisika">
                    </div>
                    <div class="form-group">
                        <label>Penerbit</label>
                        <select name="penerbit" id="penerbit_tambah">
                            <option value="">&#8212; Pilih Penerbit &#8212;</option>
                            <option value="Kemendikbud">&#127968; Kemendikbud (Resmi)</option>
                            <option value="Erlangga">Erlangga</option>
                            <option value="Intan Pariwara">Intan Pariwara</option>
                            <option value="Yudhistira">Yudhistira</option>
                            <option value="Gramedia">Gramedia</option>
                            <option value="Platinum">Platinum</option>
                            <option value="Tiga Serangkai">Tiga Serangkai</option>
                            <option value="Lainnya">Lainnya</option>
                        </select>
                    </div>
                    <div class="form-group">
                        <label>Kurikulum</label>
                        <select name="kurikulum" id="kurikulum_tambah">
                            <option value="">&#8212; Pilih Kurikulum &#8212;</option>
                            <option value="Merdeka Belajar">Merdeka Belajar</option>
                            <option value="K13">Kurikulum 2013 (K13)</option>
                            <option value="K13 Revisi">K13 Revisi</option>
                        </select>
                    </div>
                    <div class="form-group">
                        <label>Kelas / Tingkat</label>
                        <select name="kelas" id="kelas_tambah">
                            <option value="">&#8212; Pilih Kelas &#8212;</option>
                        </select>
                    </div>
                    <div class="form-group">
                        <label>Semester</label>
                        <select name="semester" id="semester_tambah">
                            <option value="">&#8212; Pilih Semester &#8212;</option>
                            <option value="Ganjil">Ganjil</option>
                            <option value="Genap">Genap</option>
                        </select>
                    </div>
                </div>
            </div>

            <div class="form-actions" style="margin-top:24px;">
                <button type="submit" class="btn btn-primary">Simpan Produk</button>
                <a href="/produk" class="btn btn-outline">Batal</a>
            </div>
        </form>
    </div>
    """
    extra = f'''<script>
    var KATEGORI_JENJANG = {_json.dumps(kategori_jenjang_map)};
    var KELAS_MAP = {{
        "SD"               : ["Kelas 1","Kelas 2","Kelas 3","Kelas 4","Kelas 5","Kelas 6"],
        "SMP"              : ["Kelas 7","Kelas 8","Kelas 9"],
        "SMA"              : ["Kelas 10","Kelas 11","Kelas 12"],
        "SMK"              : ["Kelas 10","Kelas 11","Kelas 12"],
        "Perguruan Tinggi" : ["Semester 1","Semester 2","Semester 3","Semester 4",
                              "Semester 5","Semester 6","Semester 7","Semester 8"]
    }};
    function toggleBukuFields(sel, suffix) {{
        var val  = sel.value;
        var nama = KATEGORI_JENJANG[val] || "";
        document.getElementById("buku-fields-" + suffix).style.display = nama ? "block" : "none";
        var kelasEl = document.getElementById("kelas_" + suffix);
        var opts    = KELAS_MAP[nama] || [];
        kelasEl.innerHTML = "<option value=''>\u2014 Pilih Kelas \u2014</option>";
        opts.forEach(function(o) {{
            kelasEl.innerHTML += "<option value='" + o + "'>" + o + "</option>";
        }});
    }}
    function previewFile(input, suffix) {{
        var file = input.files[0];
        if (!file) return;
        var reader = new FileReader();
        reader.onload = function(e) {{
            document.getElementById('preview-img-' + suffix).src = e.target.result;
            document.getElementById('preview-name-' + suffix).textContent = file.name;
            document.getElementById('preview-' + suffix).style.display = 'block';
            document.getElementById('drop-placeholder-' + suffix).style.display = 'none';
        }};
        reader.readAsDataURL(file);
    }}
    function handleDrop(event, suffix) {{
        event.preventDefault();
        var dz = document.getElementById('drop-zone-' + suffix);
        dz.style.borderColor = '';
        dz.style.background  = '';
        var files = event.dataTransfer.files;
        if (files.length > 0) {{
            var input = document.getElementById('gambar_' + suffix);
            // Transfer file ke input
            var dt = new DataTransfer();
            dt.items.add(files[0]);
            input.files = dt.files;
            previewFile(input, suffix);
        }}
    }}
    </script>'''
    return render_page('Tambah Produk', 'produk', content, extra)


@app.route('/produk/tambah-foto')
@login_required
def produk_tambah_foto_redirect():
    """[PYTHON] Redirect /produk/tambah-foto ke /produk/tambah yang sudah memiliki fitur upload foto."""
    return redirect('/produk/tambah')


@app.route('/produk/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def produk_edit(id):
    """[PYTHON] Edit data produk yang sudah ada, termasuk atribut buku jenjang pendidikan."""
    cur = mysql.connection.cursor()

    if request.method == 'POST':
        nama         = request.form['nama']
        kategori_id  = request.form['kategori_id'] or None
        # Hapus titik dan koma dari input angka untuk menghindari bug format desimal browser
        harga        = request.form['harga'].replace('.', '').replace(',', '')
        stok         = request.form['stok'].replace('.', '').replace(',', '') if request.form['stok'] else 0
        stok_minimum = request.form['stok_minimum'].replace('.', '').replace(',', '') if request.form['stok_minimum'] else 0
        satuan       = request.form['satuan']
        deskripsi    = request.form.get('deskripsi', '')
        status       = request.form['status']
        # [PYTHON] Atribut buku — disimpan None jika kosong
        mapel        = request.form.get('mapel', '') or None
        penerbit     = request.form.get('penerbit', '') or None
        kurikulum    = request.form.get('kurikulum', '') or None
        kelas        = request.form.get('kelas', '') or None
        semester     = request.form.get('semester', '') or None

        # [PYTHON] Proses upload gambar baru jika ada
        gambar_baru = None
        if 'gambar' in request.files:
            file = request.files['gambar']
            if file and file.filename:
                gambar_baru = save_gambar(file) # type: ignore

        # [SQL] Tambahkan kolom secara otomatis jika belum ada di database
        for col in [
            "gambar VARCHAR(255) NULL",
            "mapel VARCHAR(100) NULL",
            "penerbit VARCHAR(100) NULL",
            "kurikulum VARCHAR(100) NULL",
            "kelas VARCHAR(50) NULL",
            "semester VARCHAR(50) NULL"
        ]:
            try:
                cur.execute(f"ALTER TABLE produk ADD COLUMN {col}")
                mysql.connection.commit()
            except:
                pass

        # [SQL] UPDATE data produk beserta atribut buku
        if gambar_baru:
            cur.execute("""
                UPDATE produk
                SET nama=%s, kategori_id=%s, harga=%s, stok=%s,
                    stok_minimum=%s, satuan=%s, deskripsi=%s, status=%s, gambar=%s,
                    mapel=%s, penerbit=%s, kurikulum=%s, kelas=%s, semester=%s
                WHERE id=%s
            """, (nama, kategori_id, harga, stok, stok_minimum, satuan,
                   deskripsi, status, gambar_baru,
                   mapel, penerbit, kurikulum, kelas, semester, id))
        else:
            cur.execute("""
                UPDATE produk
                SET nama=%s, kategori_id=%s, harga=%s, stok=%s,
                    stok_minimum=%s, satuan=%s, deskripsi=%s, status=%s,
                    mapel=%s, penerbit=%s, kurikulum=%s, kelas=%s, semester=%s
                WHERE id=%s
            """, (nama, kategori_id, harga, stok, stok_minimum, satuan,
                   deskripsi, status,
                   mapel, penerbit, kurikulum, kelas, semester, id))
        mysql.connection.commit()
        cur.close()
        flash('Produk berhasil diperbarui!', 'success')
        return redirect('/produk')

    cur.execute("SELECT * FROM produk WHERE id=%s", (id,))
    data = cur.fetchone()
    cur.execute("SELECT * FROM kategori ORDER BY id")
    kategori = cur.fetchall()
    cur.close()

    if not data:
        flash('Produk tidak ditemukan!', 'danger')
        return redirect('/produk')

    # [PYTHON] Tentukan apakah produk ini kategori jenjang pendidikan
    JENJANG = ('SD', 'SMP', 'SMA', 'SMK', 'Perguruan Tinggi')
    kat_nama_aktif = next((k['nama'] for k in kategori if k['id'] == data.get('kategori_id')), '')
    is_jenjang = kat_nama_aktif in JENJANG

    # [PYTHON] Bangun opsi dropdown kategori — pisahkan jenjang pendidikan
    opt_kategori = '<option value="">&#8212; Pilih &#8212;</option>'
    opt_kategori += '<optgroup label="&#127891; Jenjang Pendidikan">'
    for k in kategori:
        if k['nama'] in JENJANG:
            sel = 'selected' if data['kategori_id'] == k['id'] else ''
            opt_kategori += f'<option value="{k["id"]}" {sel}>{k["nama"]}</option>'
    opt_kategori += '</optgroup>'
    opt_kategori += '<optgroup label="Kategori Lainnya">'
    for k in kategori:
        if k['nama'] not in JENJANG:
            sel = 'selected' if data['kategori_id'] == k['id'] else ''
            opt_kategori += f'<option value="{k["id"]}" {sel}>{k["nama"]}</option>'
    opt_kategori += '</optgroup>'

    # [PYTHON] Bangun opsi satuan
    satuans = ['pcs', 'buku', 'unit', 'kg', 'liter', 'box', 'lusin']
    opt_satuan = ''.join([f'<option value="{s}" {"selected" if data["satuan"]==s else ""}>{s}</option>' for s in satuans])

    opt_aktif    = 'selected' if data['status'] == 'aktif' else ''
    opt_nonaktif = 'selected' if data['status'] == 'tidak_aktif' else ''

    # [PYTHON] Opsi penerbit, kurikulum, kelas, semester dengan selected
    PENERBIT_LIST  = ['Kemendikbud', 'Erlangga', 'Intan Pariwara', 'Yudhistira', 'Gramedia', 'Platinum', 'Tiga Serangkai', 'Lainnya']
    KURIKULUM_LIST = [('Merdeka Belajar','Merdeka Belajar'), ('K13','Kurikulum 2013 (K13)'), ('K13 Revisi','K13 Revisi')]
    KELAS_MAP_PY   = {
        'SD': ['Kelas 1','Kelas 2','Kelas 3','Kelas 4','Kelas 5','Kelas 6'],
        'SMP': ['Kelas 7','Kelas 8','Kelas 9'],
        'SMA': ['Kelas 10','Kelas 11','Kelas 12'],
        'SMK': ['Kelas 10','Kelas 11','Kelas 12'],
        'Perguruan Tinggi': ['Semester 1','Semester 2','Semester 3','Semester 4','Semester 5','Semester 6','Semester 7','Semester 8'],
    }

    def opt_select(items, current, empty_label='— Pilih —'):
        html = f'<option value="">{empty_label}</option>'
        for item in items:
            if isinstance(item, tuple):
                val, label = item
            else:
                val = label = item
            sel = 'selected' if current == val else ''
            html += f'<option value="{val}" {sel}>{label}</option>'
        return html

    opt_penerbit  = opt_select(PENERBIT_LIST,  data.get('penerbit','') or '', '&#8212; Pilih Penerbit &#8212;')
    opt_kurikulum = opt_select(KURIKULUM_LIST, data.get('kurikulum','') or '', '&#8212; Pilih Kurikulum &#8212;')
    kelas_list    = KELAS_MAP_PY.get(kat_nama_aktif, [])
    opt_kelas     = opt_select(kelas_list, data.get('kelas','') or '', '&#8212; Pilih Kelas &#8212;')
    opt_semester  = opt_select(['Ganjil','Genap'], data.get('semester','') or '', '&#8212; Pilih Semester &#8212;')

    buku_display = 'block' if is_jenjang else 'none'

    content = f"""
    <div class="page-header">
        <div><h2>Edit Produk</h2>
        <div class="breadcrumb"><a href="/produk">Produk</a> / Edit &mdash; {data['kode_produk']}</div></div>
    </div>
    <div class="form-card" style="max-width:720px">
        <div class="form-section-title">Informasi Produk</div>
        <!-- [HTML] enctype multipart/form-data untuk support upload gambar -->
        <form method="POST" enctype="multipart/form-data">
            <div class="form-grid">
                <!-- [HTML] Preview & Upload foto produk -->
                <div class="form-group full">
                    <label>Foto Produk</label>
                    <div style="display:flex;align-items:center;gap:16px">
                        <div id="imgPreview" style="width:100px;height:100px;border-radius:8px;
                             border:1px solid var(--border);overflow:hidden;
                             background:var(--surface-2);display:flex;align-items:center;
                             justify-content:center;font-size:32px;flex-shrink:0">
                            {'<img src="/static/uploads/produk/' + data["gambar"] + '" style="width:100%;height:100%;object-fit:cover">' if data.get("gambar") else "&#128218;"}
                        </div>
                        <div>
                            <input type="file" name="gambar" id="inputGambar"
                                   accept="image/*" style="font-size:12px"
                                   onchange="previewEdit(this)">
                            <p style="font-size:11px;color:var(--text-3);margin-top:4px">
                                Kosongkan jika tidak ingin ganti foto
                            </p>
                        </div>
                    </div>
                </div>
                <div class="form-group full">
                    <label>Nama Produk *</label>
                    <input type="text" name="nama" value="{data['nama']}" required>
                </div>
                <div class="form-group">
                    <label>Kategori</label>
                    <select name="kategori_id" id="kategori_id_edit" onchange="toggleBukuFields(this, 'edit')">{opt_kategori}</select>
                </div>
                <div class="form-group"><label>Satuan</label><select name="satuan">{opt_satuan}</select></div>
                <div class="form-group">
                    <label>Harga (Rp) *</label>
                    <input type="number" name="harga" value="{int(data['harga'])}" required min="0">
                </div>
                <div class="form-group">
                    <label>Stok</label>
                    <input type="number" name="stok" value="{int(data['stok'])}" min="0">
                </div>
                <div class="form-group">
                    <label>Stok Minimum</label>
                    <input type="number" name="stok_minimum" value="{int(data['stok_minimum'])}" min="0">
                </div>
                <div class="form-group">
                    <label>Status</label>
                    <select name="status">
                        <option value="aktif" {opt_aktif}>Aktif</option>
                        <option value="tidak_aktif" {opt_nonaktif}>Tidak Aktif</option>
                    </select>
                </div>
                <div class="form-group full">
                    <label>Deskripsi</label>
                    <textarea name="deskripsi">{data['deskripsi'] or ''}</textarea>
                </div>
            </div>

            <!-- ============================================================
                 SECTION INFORMASI BUKU — tampil hanya untuk kategori jenjang pendidikan
                 ============================================================ -->
            <div id="buku-fields-edit" style="display:{buku_display}; margin-top:24px; padding-top:20px; border-top:2px solid #e2f0ff;">
                <div style="display:flex;align-items:center;gap:8px;margin-bottom:16px;">
                    <span style="font-size:20px;">&#128218;</span>
                    <div class="form-section-title" style="margin:0;">Informasi Buku</div>
                    <span style="font-size:11px;color:var(--text-3);margin-left:4px;">(opsional)</span>
                </div>
                <div class="form-grid">
                    <div class="form-group">
                        <label>Mata Pelajaran</label>
                        <input type="text" name="mapel" id="mapel_edit"
                               value="{data.get('mapel') or ''}"
                               placeholder="Misal: Matematika, Fisika, B. Indonesia">
                    </div>
                    <div class="form-group">
                        <label>Penerbit</label>
                        <select name="penerbit" id="penerbit_edit">{opt_penerbit}</select>
                    </div>
                    <div class="form-group">
                        <label>Kurikulum</label>
                        <select name="kurikulum" id="kurikulum_edit">{opt_kurikulum}</select>
                    </div>
                    <div class="form-group">
                        <label>Kelas / Tingkat</label>
                        <select name="kelas" id="kelas_edit">{opt_kelas}</select>
                    </div>
                    <div class="form-group">
                        <label>Semester</label>
                        <select name="semester" id="semester_edit">{opt_semester}</select>
                    </div>
                </div>
            </div>

            <div class="form-actions" style="margin-top:24px;">
                <button type="submit" class="btn btn-primary">Simpan Perubahan</button>
                <a href="/produk" class="btn btn-outline">Batal</a>
            </div>
        </form>
    </div>
    """
    import json as _json
    JENJANG_NAMA = ('SD', 'SMP', 'SMA', 'SMK', 'Perguruan Tinggi')
    kategori_jenjang_map = {str(k['id']): k['nama'] for k in kategori if k['nama'] in JENJANG_NAMA}
    extra = f'''<script>
    var KATEGORI_JENJANG = {_json.dumps(kategori_jenjang_map)};
    var KELAS_MAP = {{
        "SD"               : ["Kelas 1","Kelas 2","Kelas 3","Kelas 4","Kelas 5","Kelas 6"],
        "SMP"              : ["Kelas 7","Kelas 8","Kelas 9"],
        "SMA"              : ["Kelas 10","Kelas 11","Kelas 12"],
        "SMK"              : ["Kelas 10","Kelas 11","Kelas 12"],
        "Perguruan Tinggi" : ["Semester 1","Semester 2","Semester 3","Semester 4",
                              "Semester 5","Semester 6","Semester 7","Semester 8"]
    }};

    function toggleBukuFields(sel, suffix) {{
        var val    = sel.value;
        var nama   = KATEGORI_JENJANG[val] || "";
        var box    = document.getElementById("buku-fields-" + suffix);
        box.style.display = nama ? "block" : "none";
        var kelasEl = document.getElementById("kelas_" + suffix);
        var opts    = KELAS_MAP[nama] || [];
        kelasEl.innerHTML = "<option value=\'\'>\u2014 Pilih Kelas \u2014</option>";
        opts.forEach(function(o) {{
            kelasEl.innerHTML += "<option value=\'" + o + "\'>" + o + "</option>";
        }});
    }}

    function previewEdit(input) {{
        if (input.files && input.files[0]) {{
            var reader = new FileReader();
            reader.onload = function(e) {{
                document.getElementById("imgPreview").innerHTML =
                    "<img src='" + e.target.result + "' style='width:100%;height:100%;object-fit:cover'>";
            }};
            reader.readAsDataURL(input.files[0]);
        }}
    }}
    </script>'''
    return render_page('Edit Produk', 'produk', content, extra)


@app.route('/produk/hapus/<int:id>')
@login_required
def produk_hapus(id):
    """[PYTHON] Menghapus produk secara permanen dari database."""
    cur = mysql.connection.cursor()
    cur.execute("DELETE FROM produk WHERE id=%s", (id,))
    mysql.connection.commit()
    cur.close()
    flash('Produk berhasil dihapus secara permanen!', 'success')
    return redirect('/produk')


# =============================================================================
# [PYTHON] -- ROUTE PENGGUNA (monitoring user yang daftar via marketplace)
# =============================================================================

@app.route('/pengguna')
@login_required
def pengguna_index():
    """[PYTHON] Daftar pengguna/pelanggan terdaftar dari marketplace user."""
    cur = mysql.connection.cursor()
    search = request.args.get('search', '')

    query = """
        SELECT p.*,
               COUNT(DISTINCT ps.id) as total_pesanan,
               COALESCE(SUM(ps.total_bayar), 0) as total_belanja
        FROM pelanggan p
        LEFT JOIN pesanan ps ON ps.pelanggan_id = p.id
        WHERE 1=1
    """
    params = []
    if search:
        query += " AND (p.nama LIKE %s OR p.email LIKE %s)"
        params += [f'%{search}%', f'%{search}%']
    query += " GROUP BY p.id ORDER BY p.id DESC"

    cur.execute(query, params)
    pengguna = cur.fetchall()

    # Statistik ringkas
    cur.execute("SELECT COUNT(*) as c FROM pelanggan")
    total_pengguna = cur.fetchone()['c']

    cur.execute("SELECT COUNT(*) as c FROM pelanggan WHERE google_id IS NOT NULL AND google_id != ''")
    total_google = cur.fetchone()['c']

    cur.execute("""
        SELECT COUNT(DISTINCT pelanggan_id) as c FROM pesanan
    """)
    total_pernah_beli = cur.fetchone()['c']
    cur.close()

    return render_template(
        'admin/pengguna_index.html',
        pengguna=pengguna,
        total_pengguna=total_pengguna,
        total_google=total_google,
        total_pernah_beli=total_pernah_beli,
        search=search,
        page_title='Pengguna',
        active_menu='pengguna',
        now=__import__('datetime').datetime.now()
    )


# =============================================================================
# [PYTHON] -- ROUTE PENGATURAN & PERSISTENSI ADMIN
# =============================================================================

ADMIN_CONFIG_FILE = os.path.join(BASE_DIR, 'admin_config.json')

def load_admin_config():
    if os.path.exists(ADMIN_CONFIG_FILE):
        try:
            import json
            with open(ADMIN_CONFIG_FILE, 'r') as f:
                return json.load(f)
        except Exception:
            pass
    return {}

@app.before_request
def restore_admin_session():
    if 'admin_name' not in session:
        config = load_admin_config()
        if config:
            session['admin_name'] = config.get('admin_name', 'Master Admin')
            if 'admin_foto' in config:
                session['admin_foto'] = config.get('admin_foto')

@app.route('/pengaturan', methods=['GET', 'POST'])
@login_required
def pengaturan():
    """[PYTHON] Halaman pengaturan panel admin."""
    if request.method == 'POST':
        aksi = request.form.get('aksi')
        
        if aksi == 'ubah_profil':
            nama = request.form.get('nama')
            if nama:
                session['admin_name'] = nama
            
            # Proses foto profil jika ada
            if 'foto_profil' in request.files:
                file = request.files['foto_profil']
                if file and file.filename:
                    # Simpan foto pake save_gambar yang sudah ada
                    filename = save_gambar(file)
                    if filename:
                        session['admin_foto'] = f'/static/uploads/produk/{filename}'
            
            # Simpan secara permanen ke file config
            import json
            config_data = {
                'admin_name': session.get('admin_name', 'Master Admin'),
                'admin_foto': session.get('admin_foto', '')
            }
            with open(ADMIN_CONFIG_FILE, 'w') as f:
                json.dump(config_data, f)
                        
            flash('Profil admin berhasil diperbarui!', 'success')
            
        elif aksi == 'ubah_password':
            # Simulasi ubah password karena login sistem dinonaktifkan
            flash('Password berhasil diubah!', 'success')
            
        else:
            flash('Pengaturan berhasil disimpan.', 'success')
            
        return redirect('/pengaturan')
        
    return render_template(
        'admin/pengaturan.html',
        page_title='Pengaturan',
        active_menu='pengaturan',
        now=__import__('datetime').datetime.now()
    )


# =============================================================================
# [PYTHON] -- ROUTE TAMBAHAN (dipindahkan dari temp_head.py)
# =============================================================================

@app.route('/api/produk/<int:id>')
def api_produk(id):
    """
    [PYTHON] API endpoint: mengembalikan data produk dalam format JSON.
    Dipakai oleh JavaScript di halaman tambah pesanan untuk mengisi harga otomatis.
    """
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM produk WHERE id=%s", (id,))
    data = cur.fetchone()
    cur.close()
    return jsonify(data)  # jsonify: mengubah dict Python ke respons JSON


# =============================================================================
# [PYTHON] ΓöÇΓöÇ BAGIAN 10: ROUTE PESANAN (CRUD)
# =============================================================================

@app.route('/pesanan')
def pesanan_index():
    """[PYTHON] Daftar pesanan dengan filter status dan pencarian."""
    cur = mysql.connection.cursor()
    search        = request.args.get('search', '')
    status_filter = request.args.get('status', '')

    # [SQL] Query dinamis: bangun kondisi WHERE berdasarkan filter yang aktif
    query  = """
        SELECT p.*, pl.nama as nama_pelanggan
        FROM pesanan p JOIN pelanggan pl ON p.pelanggan_id = pl.id
        WHERE 1=1
    """  # WHERE 1=1: trik agar mudah menambah kondisi AND di belakangnya
    params = []

    if search:
        query += " AND (p.no_pesanan LIKE %s OR pl.nama LIKE %s)"
        params += [f'%{search}%', f'%{search}%']

    if status_filter:
        query += " AND p.status = %s"
        params.append(status_filter)

    query += " ORDER BY p.id DESC"
    cur.execute(query, params)
    data = cur.fetchall()
    cur.close()

    badge_map = {
        'pending': 'badge-gray', 'diproses': 'badge-blue',
        'dikirim': 'badge-yellow', 'selesai': 'badge-green', 'dibatalkan': 'badge-red'
    }

    rows = ''
    for p in data:
        badge = badge_map.get(p['status'], 'badge-gray')
        total = f"Rp {int(p['total_bayar']):,}".replace(',', '.')
        tgl   = p['tanggal_pesan'].strftime('%d/%m/%Y') if p['tanggal_pesan'] else '-'
        rows += f"""
        <tr>
            <td class="td-mono">{p['no_pesanan']}</td>
            <td><strong>{p['nama_pelanggan']}</strong></td>
            <td class="td-mono">{tgl}</td>
            <td class="td-mono"><strong>{total}</strong></td>
            <td class="td-mono">{p['diskon']}%</td>
            <td><span class="badge {badge}">{p['status'].upper()}</span></td>
            <td><a href="/pesanan/detail/{p['id']}" class="btn btn-outline btn-xs">Detail</a></td>
        </tr>"""

    if not rows:
        rows = '<tr><td colspan="7"><div class="empty-state"><div class="empty-icon">ΓùÄ</div><p>Belum ada pesanan</p></div></td></tr>'

    # [HTML] Dropdown filter status
    status_list = ['pending', 'diproses', 'dikirim', 'selesai', 'dibatalkan']
    opt_status = '<option value="">Semua Status</option>'
    for s in status_list:
        sel = 'selected' if status_filter == s else ''
        opt_status += f'<option value="{s}" {sel}>{s.upper()}</option>'

    content = f"""
    <div class="page-header">
        <div><h2>Daftar Pesanan</h2><div class="breadcrumb">Penjualan / Pesanan</div></div>
        <a href="/pesanan/tambah" class="btn btn-primary">+ Buat Pesanan</a>
    </div>
    <div class="table-wrap">
        <div class="table-toolbar">
            <!-- [HTML] Filter status ΓÇö onchange agar langsung submit saat dipilih -->
            <form method="GET" style="display:flex;gap:8px">
                <select name="status" onchange="this.form.submit()"
                        style="font-size:12px;padding:6px 10px;border:1px solid var(--border);border-radius:4px;font-family:inherit">
                    {opt_status}
                </select>
            </form>
            <form method="GET" style="display:flex;gap:8px;align-items:center">
                <input class="search-input" type="text" name="search"
                       placeholder="Cari no pesanan / pelanggan..." value="{search}">
                <button type="submit" class="btn btn-outline btn-sm">Cari</button>
                {'<a href="/pesanan" class="btn btn-outline btn-sm">Reset</a>' if search or status_filter else ''}
            </form>
        </div>
        <div style="overflow-x:auto">
            <table>
                <thead>
                    <tr>
                        <th>No. Pesanan</th><th>Pelanggan</th><th>Tgl. Pesan</th>
                        <th>Total Bayar</th><th>Diskon</th><th>Status</th><th>Aksi</th>
                    </tr>
                </thead>
                <tbody>{rows}</tbody>
            </table>
        </div>
    </div>
    """
    return render_page('Pesanan', 'pesanan', content)


@app.route('/pesanan/tambah', methods=['GET', 'POST'])
def pesanan_tambah():
    """[PYTHON] Membuat pesanan baru dengan banyak item produk secara dinamis."""
    cur = mysql.connection.cursor()

    if request.method == 'POST':
        no_pesanan   = generate_kode('ORD', 'pesanan', 'no_pesanan')
        pelanggan_id = request.form['pelanggan_id']
        tanggal      = request.form['tanggal_pesan']
        diskon       = float(request.form.get('diskon', 0))
        catatan      = request.form.get('catatan', '')

        # [PYTHON] getlist(): mengambil semua nilai field dengan nama yang sama (array dari form)
        # Dipakai untuk input dinamis produk_id[] dan jumlah[]
        produk_ids  = request.form.getlist('produk_id[]')
        jumlah_list = request.form.getlist('jumlah[]')

        total_harga = 0
        items       = []  # list untuk menyimpan detail item

        # [PYTHON] Iterasi setiap item yang dipilih di form
        for i in range(len(produk_ids)):
            if produk_ids[i]:  # skip jika produk tidak dipilih
                # [SQL] Ambil harga produk dari database
                cur.execute("SELECT harga, stok FROM produk WHERE id=%s", (produk_ids[i],))
                prod    = cur.fetchone()
                jumlah  = int(jumlah_list[i])
                harga_float = float(prod["harga"])  # konversi Decimal->float
                subtotal = harga_float * jumlah    # hitung subtotal per item
                total_harga += subtotal              # akumulasi total
                items.append((produk_ids[i], jumlah, harga_float, subtotal))
        total_bayar = total_harga * (1 - diskon / 100)  # total setelah diskon

        # [SQL] INSERT header pesanan
        cur.execute("""
            INSERT INTO pesanan (no_pesanan, pelanggan_id, tanggal_pesan, diskon, total_harga, total_bayar, catatan)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (no_pesanan, pelanggan_id, tanggal, diskon, total_harga, total_bayar, catatan))

        pesanan_id = cur.lastrowid  # lastrowid: id yang baru saja di-INSERT

        # [SQL] INSERT detail pesanan dan update stok produk
        for item in items:
            cur.execute("""
                INSERT INTO detail_pesanan (pesanan_id, produk_id, jumlah, harga_satuan, subtotal)
                VALUES (%s, %s, %s, %s, %s)
            """, (pesanan_id, item[0], item[1], item[2], item[3]))

            # [SQL] Kurangi stok produk setelah pesanan dibuat
            cur.execute("UPDATE produk SET stok = stok - %s WHERE id=%s", (item[1], item[0]))

        mysql.connection.commit()
        cur.close()
        flash(f'Pesanan {no_pesanan} berhasil dibuat!', 'success')
        return redirect('/pesanan')

    # [SQL] Data untuk dropdown pelanggan dan produk
    cur.execute("SELECT id, kode_pelanggan, nama FROM pelanggan WHERE status='aktif'")
    pelanggan = cur.fetchall()

    cur.execute("SELECT id, kode_produk, nama, harga, stok, satuan FROM produk WHERE status='aktif' AND stok > 0")
    produk = cur.fetchall()
    cur.close()

    # [PYTHON] Bangun opsi dropdown pelanggan
    opt_pelanggan = '<option value="">ΓÇö Pilih Pelanggan ΓÇö</option>'
    for p in pelanggan:
        opt_pelanggan += f'<option value="{p["id"]}">[{p["kode_pelanggan"]}] {p["nama"]}</option>'

    # [PYTHON] Bangun opsi dropdown produk dengan data- attribute untuk harga dan stok
    opt_produk = '<option value="">ΓÇö Pilih Produk ΓÇö</option>'
    for p in produk:
        harga = f"Rp {int(p['harga']):,}".replace(',', '.')
        # [HTML] data-harga dan data-stok: custom attribute untuk diakses JavaScript
        opt_produk += f'<option value="{p["id"]}" data-harga="{p["harga"]}" data-stok="{p["stok"]}">[{p["kode_produk"]}] {p["nama"]} ΓÇö {harga} (Stok: {p["stok"]} {p["satuan"]})</option>'

    today = date.today().isoformat()  # format tanggal: YYYY-MM-DD

    content = f"""
    <div class="page-header">
        <div><h2>Buat Pesanan Baru</h2>
        <div class="breadcrumb"><a href="/pesanan">Pesanan</a> / Buat Baru</div></div>
    </div>

    <!-- [HTML] Template tersembunyi: opsi produk untuk di-clone saat tambah baris -->
    <template id="produk-options-template">{opt_produk}</template>

    <form method="POST">
        <div class="grid-2" style="align-items:start">

            <!-- Kolom kiri: info pesanan + ringkasan -->
            <div>
                <div class="form-card" style="margin-bottom:16px">
                    <div class="form-section-title">Informasi Pesanan</div>
                    <div class="form-grid">
                        <div class="form-group">
                            <label>Pelanggan *</label>
                            <select name="pelanggan_id" required>{opt_pelanggan}</select>
                        </div>
                        <div class="form-group">
                            <label>Tanggal Pesanan *</label>
                            <!-- [HTML] type="date" ΓåÆ input kalender bawaan browser -->
                            <input type="date" name="tanggal_pesan" value="{today}" required>
                        </div>
                        <div class="form-group">
                            <label>Diskon (%)</label>
                            <!-- [HTML] oninput: event saat user mengetik, langsung hitung ulang total -->
                            <input type="number" name="diskon" id="diskon" value="0"
                                   min="0" max="100" step="0.1" oninput="hitungTotal()">
                        </div>
                        <div class="form-group full">
                            <label>Catatan</label>
                            <textarea name="catatan" placeholder="Opsional" style="min-height:60px"></textarea>
                        </div>
                    </div>
                </div>

                <!-- [HTML] Kotak ringkasan total harga -->
                <div class="summary-box">
                    <div class="summary-row">
                        <span style="color:var(--text-2)">Total Harga</span>
                        <!-- [HTML] id dipakai JavaScript untuk update nilai secara realtime -->
                        <span id="total-harga-display">Rp 0</span>
                    </div>
                    <div class="summary-row">
                        <span style="color:var(--text-2)">Diskon</span>
                        <span id="diskon-nominal" style="color:var(--red)">Rp 0</span>
                    </div>
                    <div class="summary-row">
                        <span>Total Bayar</span>
                        <span id="total-bayar-display" style="color:var(--green)">Rp 0</span>
                    </div>
                </div>

                <div style="display:flex;gap:10px;margin-top:16px">
                    <button type="submit" class="btn btn-primary">Γ£ô Buat Pesanan</button>
                    <a href="/pesanan" class="btn btn-outline">Batal</a>
                </div>
            </div>

            <!-- Kolom kanan: baris item produk dinamis -->
            <div class="form-card">
                <div class="form-section-title" style="display:flex;justify-content:space-between;align-items:center">
                    <span>Item Produk</span>
                    <!-- [HTML] onclick: memanggil fungsi JavaScript addItem() -->
                    <button type="button" class="btn btn-outline btn-xs" onclick="addItem()">+ Tambah Item</button>
                </div>

                <!-- [HTML] Container untuk baris item dinamis -->
                <div id="items-container">
                    <!-- Baris item pertama (statis) -->
                    <div class="item-row" id="item-row-1">
                        <div class="form-group">
                            <label>Produk</label>
                            <!-- [HTML] name="produk_id[]" ΓåÆ array, onchange ΓåÆ update harga -->
                            <select name="produk_id[]" onchange="updateHarga(this, 1)" required>
                                {opt_produk}
                            </select>
                        </div>
                        <div class="form-group">
                            <label>Jumlah</label>
                            <input type="number" name="jumlah[]" id="jumlah-1"
                                   value="1" min="1" oninput="hitungSubtotal(1)" required>
                        </div>
                        <div class="form-group">
                            <label>Harga</label>
                            <input type="text" id="harga-display-1" readonly placeholder="Rp 0">
                            <input type="hidden" id="harga-1" value="0">
                        </div>
                        <div class="form-group">
                            <label>Subtotal</label>
                            <input type="text" id="subtotal-display-1" readonly placeholder="Rp 0">
                        </div>
                        <div class="form-group">
                            <label>&nbsp;</label>
                            <!-- [HTML] visibility:hidden agar baris pertama tidak bisa dihapus -->
                            <button type="button" class="btn btn-danger btn-sm" style="visibility:hidden">Γ£ò</button>
                        </div>
                    </div>
                </div>

                <!-- [HTML] Informasi stok -->
                <div style="margin-top:8px;padding:10px;background:var(--surface-2);border-radius:4px;font-size:12px;color:var(--text-3)">
                    Γôÿ Stok akan otomatis berkurang setelah pesanan dibuat
                </div>
            </div>

        </div>
    </form>
    """
    return render_page('Buat Pesanan', 'pesanan', content)


@app.route('/pesanan/detail/<int:id>')
def pesanan_detail(id):
    """[PYTHON] Menampilkan detail pesanan dan form update status."""
    cur = mysql.connection.cursor()

    # [SQL] JOIN 3 tabel: pesanan + pelanggan untuk data lengkap
    cur.execute("""
        SELECT p.*, pl.nama as nama_pelanggan, pl.telepon, pl.alamat, pl.kota
        FROM pesanan p JOIN pelanggan pl ON p.pelanggan_id = pl.id
        WHERE p.id = %s
    """, (id,))
    pesanan = cur.fetchone()

    # [SQL] JOIN detail_pesanan + produk untuk item-item dalam pesanan
    cur.execute("""
        SELECT dp.*, pr.nama as nama_produk, pr.kode_produk, pr.satuan
        FROM detail_pesanan dp JOIN produk pr ON dp.produk_id = pr.id
        WHERE dp.pesanan_id = %s
    """, (id,))
    detail = cur.fetchall()
    cur.close()

    if not pesanan:
        flash('Pesanan tidak ditemukan!', 'danger')
        return redirect('/pesanan')

    badge_map = {
        'pending': 'badge-gray', 'diproses': 'badge-blue',
        'dikirim': 'badge-yellow', 'selesai': 'badge-green', 'dibatalkan': 'badge-red'
    }
    badge = badge_map.get(pesanan['status'], 'badge-gray')

    # [PYTHON] Format tanggal untuk tampilan
    tgl_pesan = pesanan['tanggal_pesan'].strftime('%d %B %Y') if pesanan['tanggal_pesan'] else '-'
    tgl_kirim = pesanan['tanggal_kirim'].strftime('%d %B %Y') if pesanan['tanggal_kirim'] else 'ΓÇö'
    tgl_kirim_val = pesanan['tanggal_kirim'].isoformat() if pesanan['tanggal_kirim'] else ''

    total_harga = f"Rp {int(pesanan['total_harga']):,}".replace(',', '.')
    total_bayar = f"Rp {int(pesanan['total_bayar']):,}".replace(',', '.')

    # [PYTHON] Bangun baris item detail
    item_rows = ''
    for item in detail:
        harga_sat = f"Rp {int(item['harga_satuan']):,}".replace(',', '.')
        subtotal  = f"Rp {int(item['subtotal']):,}".replace(',', '.')
        item_rows += f"""
        <tr>
            <td><strong>{item['nama_produk']}</strong>
                <div class="td-mono" style="font-size:11px">{item['kode_produk']}</div>
            </td>
            <td class="td-mono">{item['jumlah']} {item['satuan']}</td>
            <td class="td-mono">{harga_sat}</td>
            <td class="td-mono"><strong>{subtotal}</strong></td>
        </tr>"""

    # [HTML] Dropdown status untuk form update
    status_list = ['pending', 'diproses', 'dikirim', 'selesai', 'dibatalkan']
    opt_status = ''
    for s in status_list:
        sel = 'selected' if pesanan['status'] == s else ''
        opt_status += f'<option value="{s}" {sel}>{s.upper()}</option>'

    content = f"""
    <div class="page-header">
        <div><h2>Detail Pesanan</h2>
        <div class="breadcrumb"><a href="/pesanan">Pesanan</a> / {pesanan['no_pesanan']}</div></div>
        <span class="badge {badge}" style="font-size:13px;padding:5px 12px">{pesanan['status'].upper()}</span>
    </div>

    <div class="grid-2" style="align-items:start">
        <!-- Informasi pesanan dan pelanggan -->
        <div>
            <div class="card" style="margin-bottom:16px">
                <div class="card-header"><span class="card-title">Informasi Pesanan</span></div>
                <div class="card-body">
                    <div class="info-grid">
                        <div class="info-row">
                            <span class="info-label">No. Pesanan</span>
                            <span class="info-value td-mono">{pesanan['no_pesanan']}</span>
                        </div>
                        <div class="info-row">
                            <span class="info-label">Tanggal Pesan</span>
                            <span class="info-value">{tgl_pesan}</span>
                        </div>
                        <div class="info-row">
                            <span class="info-label">Tanggal Kirim</span>
                            <span class="info-value">{tgl_kirim}</span>
                        </div>
                        <div class="info-row">
                            <span class="info-label">Diskon</span>
                            <span class="info-value">{pesanan['diskon']}%</span>
                        </div>
                        <div class="info-row">
                            <span class="info-label">Total Harga</span>
                            <span class="info-value td-mono">{total_harga}</span>
                        </div>
                        <div class="info-row">
                            <span class="info-label">Total Bayar</span>
                            <span class="info-value td-mono" style="font-size:16px;color:var(--green);font-weight:700">{total_bayar}</span>
                        </div>
                    </div>
                </div>
            </div>
            <div class="card">
                <div class="card-header"><span class="card-title">Pelanggan</span></div>
                <div class="card-body">
                    <div class="info-grid">
                        <div class="info-row">
                            <span class="info-label">Nama</span>
                            <span class="info-value">{pesanan['nama_pelanggan']}</span>
                        </div>
                        <div class="info-row">
                            <span class="info-label">Telepon</span>
                            <span class="info-value td-mono">{pesanan['telepon'] or '-'}</span>
                        </div>
                        <div class="info-row" style="grid-column:1/-1">
                            <span class="info-label">Alamat</span>
                            <span class="info-value">{pesanan['alamat'] or '-'}, {pesanan['kota'] or ''}</span>
                        </div>
                    </div>
                </div>
            </div>
        </div>

        <!-- Update status + item pesanan -->
        <div>
            <div class="card" style="margin-bottom:16px">
                <div class="card-header"><span class="card-title">Update Status</span></div>
                <div class="card-body">
                    <!-- [HTML] Form update status menggunakan method POST -->
                    <form method="POST" action="/pesanan/update_status/{id}">
                        <div class="form-grid">
                            <div class="form-group">
                                <label>Status Baru</label>
                                <select name="status">{opt_status}</select>
                            </div>
                            <div class="form-group">
                                <label>Tanggal Kirim</label>
                                <input type="date" name="tanggal_kirim" value="{tgl_kirim_val}">
                            </div>
                        </div>
                        <div style="display:flex;gap:10px;margin-top:12px">
                            <button type="submit" class="btn btn-primary btn-sm">Simpan Status</button>
                            <a href="/pesanan" class="btn btn-outline btn-sm">ΓåÉ Kembali</a>
                            <a href="/pesanan/invoice/{id}" class="btn btn-primary btn-sm" target="_blank">
                                ≡ƒû¿ Cetak Invoice PDF
                            </a>
                        </div>
                    </form>
                </div>
            </div>
            <div class="card">
                <div class="card-header"><span class="card-title">Item Pesanan</span></div>
                <table>
                    <thead>
                        <tr><th>Produk</th><th>Jumlah</th><th>Harga</th><th>Subtotal</th></tr>
                    </thead>
                    <tbody>
                        {item_rows}
                        <!-- [HTML] Baris total di bagian bawah tabel -->
                        <tr style="background:var(--surface-2)">
                            <td colspan="3" style="text-align:right;font-weight:700;font-size:12px">TOTAL BAYAR</td>
                            <td class="td-mono" style="font-weight:700;color:var(--green)">{total_bayar}</td>
                        </tr>
                    </tbody>
                </table>
            </div>
        </div>
    </div>
    """
    return render_page('Detail Pesanan', 'pesanan', content)


@app.route('/pesanan/update_status/<int:id>', methods=['POST'])
def pesanan_update_status(id):
    """[PYTHON] Memperbarui status dan tanggal kirim pesanan."""
    status       = request.form['status']
    tanggal_kirim = request.form.get('tanggal_kirim')

    cur = mysql.connection.cursor()
    if tanggal_kirim:
        cur.execute("UPDATE pesanan SET status=%s, tanggal_kirim=%s WHERE id=%s",
                    (status, tanggal_kirim, id))
    else:
        cur.execute("UPDATE pesanan SET status=%s WHERE id=%s", (status, id))

    mysql.connection.commit()
    cur.close()
    flash('Status pesanan berhasil diperbarui!', 'success')
    return redirect(f'/pesanan/detail/{id}')


# =============================================================================
# [PYTHON] ΓöÇΓöÇ BAGIAN 11: ROUTE LAPORAN
# =============================================================================

@app.route('/laporan')
def laporan_index():
    """[PYTHON] Laporan penjualan bulanan: ringkasan, produk terlaris, pelanggan terbaik."""
    cur = mysql.connection.cursor()

    # [PYTHON] Ambil parameter bulan dari URL, default bulan sekarang
    bulan = request.args.get('bulan', datetime.now().strftime('%Y-%m'))
    tahun = bulan[:4]   # ambil 4 karakter pertama: tahun
    bln   = bulan[5:7]  # ambil karakter ke-5 dan 6: bulan (01-12)

    # [SQL] Ringkasan statistik bulan ini
    cur.execute("""
        SELECT COUNT(*) as total_pesanan,
               COALESCE(SUM(total_bayar), 0)  as total_pendapatan,
               COALESCE(AVG(total_bayar), 0)  as rata_rata
        FROM pesanan
        WHERE status = 'selesai'
          AND MONTH(tanggal_pesan) = %s
          AND YEAR(tanggal_pesan)  = %s
    """, (bln, tahun))
    ringkasan = cur.fetchone()

    # [SQL] 5 produk terlaris berdasarkan jumlah terjual
    # GROUP BY: mengelompokkan baris berdasarkan produk
    # ORDER BY total_terjual DESC: urutkan dari terbanyak
    cur.execute("""
        SELECT pr.nama,
               SUM(dp.jumlah)    as total_terjual,
               SUM(dp.subtotal)  as total_pendapatan
        FROM detail_pesanan dp
        JOIN produk pr   ON dp.produk_id  = pr.id
        JOIN pesanan p   ON dp.pesanan_id = p.id
        WHERE p.status = 'selesai'
          AND MONTH(p.tanggal_pesan) = %s
          AND YEAR(p.tanggal_pesan)  = %s
        GROUP BY pr.id
        ORDER BY total_terjual DESC
        LIMIT 5
    """, (bln, tahun))
    produk_terlaris = cur.fetchall()

    # [SQL] 5 pelanggan dengan total belanja terbesar
    cur.execute("""
        SELECT pl.nama, pl.kota,
               COUNT(p.id)       as total_pesanan,
               SUM(p.total_bayar) as total_belanja
        FROM pesanan p JOIN pelanggan pl ON p.pelanggan_id = pl.id
        WHERE p.status = 'selesai'
          AND MONTH(p.tanggal_pesan) = %s
          AND YEAR(p.tanggal_pesan)  = %s
        GROUP BY pl.id
        ORDER BY total_belanja DESC
        LIMIT 5
    """, (bln, tahun))
    pelanggan_terbaik = cur.fetchall()

    # [SQL] Semua pesanan di bulan ini (semua status)
    cur.execute("""
        SELECT p.no_pesanan, pl.nama as nama_pelanggan,
               p.tanggal_pesan, p.total_bayar, p.status
        FROM pesanan p JOIN pelanggan pl ON p.pelanggan_id = pl.id
        WHERE MONTH(p.tanggal_pesan) = %s AND YEAR(p.tanggal_pesan) = %s
        ORDER BY p.tanggal_pesan ASC
    """, (bln, tahun))
    semua_pesanan = cur.fetchall()
    cur.close()

    # [PYTHON] Format nilai Rupiah
    total_pend = f"Rp {int(ringkasan['total_pendapatan']):,}".replace(',', '.')
    rata_rata  = f"Rp {int(ringkasan['rata_rata']):,}".replace(',', '.')

    # [PYTHON] Bangun baris tabel produk terlaris
    produk_rows = ''
    for i, p in enumerate(produk_terlaris):
        pend = f"Rp {int(p['total_pendapatan']):,}".replace(',', '.')
        produk_rows += f"""
        <tr>
            <td style="font-weight:700;color:var(--text-3)">{i+1}</td>
            <td><strong>{p['nama']}</strong></td>
            <td class="td-mono">{p['total_terjual']}</td>
            <td class="td-mono">{pend}</td>
        </tr>"""
    if not produk_rows:
        produk_rows = '<tr><td colspan="4"><div class="empty-state"><p>Belum ada data</p></div></td></tr>'

    # [PYTHON] Bangun baris tabel pelanggan terbaik
    pelanggan_rows = ''
    for i, p in enumerate(pelanggan_terbaik):
        belanja = f"Rp {int(p['total_belanja']):,}".replace(',', '.')
        pelanggan_rows += f"""
        <tr>
            <td style="font-weight:700;color:var(--text-3)">{i+1}</td>
            <td><strong>{p['nama']}</strong><div style="font-size:11px;color:var(--text-3)">{p['kota']}</div></td>
            <td class="td-mono">{p['total_pesanan']}x</td>
            <td class="td-mono">{belanja}</td>
        </tr>"""
    if not pelanggan_rows:
        pelanggan_rows = '<tr><td colspan="4"><div class="empty-state"><p>Belum ada data</p></div></td></tr>'

    # [PYTHON] Bangun baris semua transaksi
    badge_map = {
        'pending': 'badge-gray', 'diproses': 'badge-blue',
        'dikirim': 'badge-yellow', 'selesai': 'badge-green', 'dibatalkan': 'badge-red'
    }
    transaksi_rows = ''
    for p in semua_pesanan:
        badge = badge_map.get(p['status'], 'badge-gray')
        total = f"Rp {int(p['total_bayar']):,}".replace(',', '.')
        tgl   = p['tanggal_pesan'].strftime('%d/%m/%Y') if p['tanggal_pesan'] else '-'
        transaksi_rows += f"""
        <tr>
            <td class="td-mono">{p['no_pesanan']}</td>
            <td>{p['nama_pelanggan']}</td>
            <td class="td-mono">{tgl}</td>
            <td class="td-mono"><strong>{total}</strong></td>
            <td><span class="badge {badge}">{p['status'].upper()}</span></td>
        </tr>"""
    if not transaksi_rows:
        transaksi_rows = '<tr><td colspan="5"><div class="empty-state"><div class="empty-icon">Γùº</div><p>Tidak ada transaksi periode ini</p></div></td></tr>'

    content = f"""
    <div class="page-header">
        <div><h2>Laporan Bulanan</h2><div class="breadcrumb">Analitik / Laporan</div></div>
        <!-- [HTML] Form filter periode bulan -->
        <!-- [HTML] action="/laporan" ΓåÆ form submit ke URL yang benar -->
        <form action="/laporan" method="GET" style="display:flex;gap:8px;align-items:center;flex-wrap:wrap">
            <label style="font-size:12px;color:var(--text-2);font-weight:600;letter-spacing:0.06em">PERIODE</label>
            <!-- [HTML] type="month" ΓåÆ input kalender bulan-tahun bawaan browser -->
            <input type="month" name="bulan" value="{bulan}"
                   style="font-family:inherit;font-size:13px;padding:7px 10px;
                          border:1px solid var(--border);border-radius:4px;outline:none">
            <button type="submit" class="btn btn-primary btn-sm">Tampilkan</button>
            <!-- [HTML] Tombol export ΓÇö href pakai f-string {bulan} dari Python -->
            <a href="/laporan/export?bulan={bulan}"
               style="display:inline-flex;align-items:center;gap:5px;padding:5px 12px;
                      background:#2d6a4f;color:#fff;border-radius:4px;text-decoration:none;
                      font-size:12px;font-weight:600">
                &#128202; Export Excel (Bulan Ini)
            </a>
            <a href="/laporan/export"
               style="display:inline-flex;align-items:center;gap:5px;padding:5px 12px;
                      background:#1a4a8a;color:#fff;border-radius:4px;text-decoration:none;
                      font-size:12px;font-weight:600">
                &#128190; Export Semua Data
            </a>
        </form>
    </div>

    <!-- [HTML] Statistik ringkasan bulan ini -->
    <div class="stats-grid" style="margin-bottom:24px">
        <div class="stat-card green">
            <div class="stat-label">Pesanan Selesai</div>
            <div class="stat-value">{ringkasan['total_pesanan']}</div>
            <div class="stat-sub">order berhasil</div>
        </div>
        <div class="stat-card yellow">
            <div class="stat-label">Total Pendapatan</div>
            <div class="stat-value" style="font-size:15px">{total_pend}</div>
        </div>
        <div class="stat-card blue">
            <div class="stat-label">Rata-rata per Order</div>
            <div class="stat-value" style="font-size:15px">{rata_rata}</div>
        </div>
        <div class="stat-card">
            <div class="stat-label">Periode</div>
            <div class="stat-value" style="font-size:18px;font-family:'IBM Plex Mono',monospace">{bulan}</div>
        </div>
    </div>

    <!-- [HTML] Tabel produk terlaris dan pelanggan terbaik berdampingan -->
    <div class="grid-2">
        <div class="card">
            <div class="card-header"><span class="card-title">≡ƒÅå Produk Terlaris</span></div>
            <table>
                <thead><tr><th>#</th><th>Produk</th><th>Terjual</th><th>Pendapatan</th></tr></thead>
                <tbody>{produk_rows}</tbody>
            </table>
        </div>
        <div class="card">
            <div class="card-header"><span class="card-title">Γ¡É Pelanggan Terbaik</span></div>
            <table>
                <thead><tr><th>#</th><th>Pelanggan</th><th>Order</th><th>Total Belanja</th></tr></thead>
                <tbody>{pelanggan_rows}</tbody>
            </table>
        </div>
    </div>

    <!-- [HTML] Tabel semua transaksi bulan ini -->
    <div class="table-wrap" style="margin-top:20px">
        <div class="table-toolbar">
            <span style="font-size:13px;font-weight:700">Semua Transaksi ΓÇö {bulan}</span>
            <span class="badge badge-gray">{len(semua_pesanan)} transaksi</span>
        </div>
        <div style="overflow-x:auto">
            <table>
                <thead>
                    <tr><th>No. Pesanan</th><th>Pelanggan</th><th>Tanggal</th><th>Total Bayar</th><th>Status</th></tr>
                </thead>
                <tbody>{transaksi_rows}</tbody>
            </table>
        </div>
    </div>
    """
    return render_page('Laporan', 'laporan', content)


# =============================================================================
# [PYTHON] ΓöÇΓöÇ BAGIAN 12: MENJALANKAN APLIKASI
# =============================================================================


# =============================================================================
# [PYTHON] -- BAGIAN 12B: KONFIGURASI EMAIL GMAIL
# =============================================================================
EMAIL_SENDER   = 'emailkamu@gmail.com'
EMAIL_PASSWORD = 'xxxx xxxx xxxx xxxx'
EMAIL_ADMIN    = 'emailkamu@gmail.com'

def kirim_email_stok(produk_menipis):
    if not produk_menipis: return
    try:
        baris = "".join(f'<tr><td style="padding:8px">{p["nama"]}</td><td style="padding:8px;color:red">{p["stok"]} {p["satuan"]}</td><td style="padding:8px">{p["stok_minimum"]}</td></tr>' for p in produk_menipis)
        isi = f"""<html><body style="font-family:Arial"><div style="background:#052e16;padding:20px;text-align:center"><h2 style="color:#e8c547">PendidikanStore - Peringatan Stok!</h2></div><div style="padding:20px"><table style="width:100%;border-collapse:collapse"><tr style="background:#052e16;color:#fff"><th style="padding:10px">Produk</th><th>Stok</th><th>Minimum</th></tr>{baris}</table><p>Segera restok!</p></div></body></html>"""
        msg = MIMEMultipart("alternative")
        msg["Subject"] = f"Peringatan: {len(produk_menipis)} Produk Stok Menipis - PendidikanStore"
        msg["From"] = EMAIL_SENDER
        msg["To"]   = EMAIL_ADMIN
        msg.attach(MIMEText(isi, "html"))
        with smtplib.SMTP("smtp.gmail.com", 587) as server:
            server.starttls()
            server.login(EMAIL_SENDER, EMAIL_PASSWORD)
            server.sendmail(EMAIL_SENDER, EMAIL_ADMIN, msg.as_string())
        print(f"[EMAIL OK] Terkirim ke {EMAIL_ADMIN}")
    except Exception as e:
        print(f"[EMAIL ERROR] {e}")


# =============================================================================
# [PYTHON] -- BAGIAN 13: CETAK INVOICE PDF
# =============================================================================
@app.route("/pesanan/invoice/<int:id>")
def cetak_invoice(id):
    cur = mysql.connection.cursor()
    cur.execute("""SELECT p.*, pl.nama as nama_pelanggan, pl.telepon, pl.alamat, pl.kota
        FROM pesanan p JOIN pelanggan pl ON p.pelanggan_id=pl.id WHERE p.id=%s""", (id,))
    pesanan = cur.fetchone()
    cur.execute("""SELECT dp.*, pr.nama as nama_produk, pr.kode_produk, pr.satuan
        FROM detail_pesanan dp JOIN produk pr ON dp.produk_id=pr.id
        WHERE dp.pesanan_id=%s""", (id,))
    detail = cur.fetchall()
    cur.close()
    if not pesanan:
        flash("Pesanan tidak ditemukan!", "danger")
        return redirect("/pesanan")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=20*mm, leftMargin=20*mm, topMargin=20*mm, bottomMargin=20*mm)
    styles = getSampleStyleSheet()
    def sty(**kw): return ParagraphStyle("x", parent=styles["Normal"], **kw)
    el = []

    hd = [[Paragraph("<b>PendidikanStore</b>", sty(fontSize=20, fontName="Helvetica-Bold")),
           Paragraph(f"<b>INVOICE #{pesanan['no_pesanan']}</b>", sty(fontSize=14, fontName="Helvetica-Bold", alignment=TA_RIGHT))],
          [Paragraph("Alat Tulis & Perlengkapan Sekolah - APSI Teknik Industri", sty(fontSize=9, textColor=colors.grey)),
           Paragraph(f"{pesanan['tanggal_pesan'].strftime('%d %B %Y') if pesanan['tanggal_pesan'] else '-'}", sty(fontSize=9, alignment=TA_RIGHT, textColor=colors.grey))]]
    th = Table(hd, colWidths=[95*mm, 75*mm])
    th.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"TOP")]))
    el.extend([th, HRFlowable(width="100%", thickness=2, color=colors.HexColor("#052e16"), spaceAfter=5*mm)])

    info = [[Paragraph("<b>KEPADA:</b>", sty(fontSize=9, fontName="Helvetica-Bold")), "",
             Paragraph("<b>STATUS:</b>", sty(fontSize=9, fontName="Helvetica-Bold"))],
            [Paragraph(f"<b>{pesanan['nama_pelanggan']}</b>", sty(fontSize=11, fontName="Helvetica-Bold")), "",
             Paragraph(f"<b>{pesanan['status'].upper()}</b>", sty(fontSize=10, fontName="Helvetica-Bold"))],
            [Paragraph(pesanan["alamat"] or "-", sty(fontSize=9)), "",
             Paragraph(f"Tgl. Kirim: {pesanan['tanggal_kirim'].strftime('%d %B %Y') if pesanan['tanggal_kirim'] else 'Belum dikirim'}", sty(fontSize=9))],
            [Paragraph(f"{pesanan['kota'] or ''} | {pesanan['telepon'] or ''}", sty(fontSize=9)), "", Paragraph("", sty(fontSize=9))]]
    ti = Table(info, colWidths=[88*mm, 8*mm, 74*mm])
    ti.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"TOP"),("BACKGROUND",(0,0),(0,-1),colors.HexColor("#f9f8f6")),("BACKGROUND",(2,0),(2,-1),colors.HexColor("#f9f8f6")),("LEFTPADDING",(0,0),(0,-1),8),("LEFTPADDING",(2,0),(2,-1),8),("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5)]))
    el.extend([ti, Spacer(1, 5*mm)])

    rows = [["No","Kode","Nama Produk","Sat.","Qty","Harga","Subtotal"]]
    for i, item in enumerate(detail):
        rows.append([str(i+1), item["kode_produk"], item["nama_produk"], item["satuan"], str(item["jumlah"]),
                     f"Rp {int(item['harga_satuan']):,}".replace(",","."),
                     f"Rp {int(item['subtotal']):,}".replace(",",".")])
    tbl = Table(rows, colWidths=[8*mm,20*mm,57*mm,12*mm,10*mm,28*mm,30*mm])
    tbl.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor("#052e16")),("TEXTCOLOR",(0,0),(-1,0),colors.white),("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),8),("ALIGN",(4,1),(-1,-1),"RIGHT"),("ALIGN",(0,0),(3,-1),"CENTER"),("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5),("GRID",(0,0),(-1,-1),0.5,colors.HexColor("#bbf7d0")),("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#f9f8f6")])]))
    el.extend([tbl, Spacer(1, 4*mm)])

    dn = float(pesanan["total_harga"]) * float(pesanan["diskon"]) / 100
    tot = [["","Subtotal:", f"Rp {int(pesanan['total_harga']):,}".replace(",",".")],
           ["",f"Diskon ({pesanan['diskon']}%):", f"- Rp {int(dn):,}".replace(",",".")],
           ["","TOTAL BAYAR:", f"Rp {int(pesanan['total_bayar']):,}".replace(",",".")]]
    tt = Table(tot, colWidths=[100*mm, 40*mm, 30*mm])
    tt.setStyle(TableStyle([("ALIGN",(1,0),(-1,-1),"RIGHT"),("FONTSIZE",(0,0),(-1,-1),9),("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),("LINEABOVE",(1,2),(-1,2),1.5,colors.HexColor("#052e16")),("FONTNAME",(1,2),(-1,2),"Helvetica-Bold"),("FONTSIZE",(1,2),(-1,2),11),("TEXTCOLOR",(2,2),(2,2),colors.HexColor("#2d6a4f"))]))
    el.extend([tt, Spacer(1,6*mm), HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#bbf7d0"), spaceAfter=3*mm),
               Paragraph("Terima kasih atas kepercayaan Anda! -- PendidikanStore APSI Teknik Industri", sty(fontSize=8, textColor=colors.grey, alignment=TA_CENTER))])
    doc.build(el)
    buf.seek(0)
    return send_file(buf, mimetype="application/pdf", as_attachment=True, download_name=f"Invoice_{pesanan['no_pesanan']}.pdf")


# =============================================================================
# [PYTHON] -- BAGIAN 14: EXPORT LAPORAN EXCEL (3 Sheet)
# =============================================================================
@app.route("/laporan/export")
def export_excel():
    cur = mysql.connection.cursor()
    bulan = request.args.get("bulan", "")
    tahun = bulan[:4] if bulan else ""
    bln   = bulan[5:7] if bulan else ""

    q_p = """SELECT p.no_pesanan, pl.nama as pelanggan, pl.kota, p.tanggal_pesan, p.tanggal_kirim,
               p.status, p.total_harga, p.diskon, p.total_bayar
            FROM pesanan p JOIN pelanggan pl ON p.pelanggan_id=pl.id"""
    q_s = """SELECT p.no_pesanan, p.tanggal_pesan, pl.nama as pelanggan, pr.kode_produk,
               pr.nama as produk, pr.satuan, dp.jumlah, dp.harga_satuan, dp.subtotal, p.status
            FROM detail_pesanan dp JOIN pesanan p ON dp.pesanan_id=p.id
            JOIN pelanggan pl ON p.pelanggan_id=pl.id JOIN produk pr ON dp.produk_id=pr.id"""

    if bulan:
        cur.execute(q_p + " WHERE MONTH(p.tanggal_pesan)=%s AND YEAR(p.tanggal_pesan)=%s ORDER BY p.tanggal_pesan", (bln,tahun))
    else:
        cur.execute(q_p + " ORDER BY p.tanggal_pesan ASC")
    dp = cur.fetchall()

    if bulan:
        cur.execute(q_s + " WHERE MONTH(p.tanggal_pesan)=%s AND YEAR(p.tanggal_pesan)=%s ORDER BY p.tanggal_pesan", (bln,tahun))
    else:
        cur.execute(q_s + " ORDER BY p.tanggal_pesan ASC")
    ds = cur.fetchall()

    cur.execute("""SELECT pr.kode_produk, pr.nama, k.nama as kategori, pr.harga, pr.stok,
               pr.stok_minimum, pr.satuan, (pr.stok*pr.harga) as nilai_stok
            FROM produk pr LEFT JOIN kategori k ON pr.kategori_id=k.id ORDER BY k.nama, pr.nama""")
    dr = cur.fetchall()
    cur.close()

    wb = Workbook()
    CH="FF1A1916"; CG="FF2D6A4F"; CR="FFC0392B"; CST="FFF9F8F6"; CBR="FFE2DFD8"; FMT="#,##0"

    def bd():
        s=Side(style="thin",color=CBR); return Border(left=s,right=s,top=s,bottom=s)
    def sh(cell,val):
        cell.value=val; cell.font=Font(name="Arial",bold=True,color="FFFFFFFF",size=10)
        cell.fill=PatternFill("solid",fgColor=CH); cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); cell.border=bd()
    def sd(cell,val,al="left",bold=False,bg=None,col=None):
        cell.value=val; cell.font=Font(name="Arial",bold=bold,size=9,color=col if col else "FF1A1916")
        cell.alignment=Alignment(horizontal=al,vertical="center"); cell.border=bd()
        if bg: cell.fill=PatternFill("solid",fgColor=bg)
    def st(cell,val):
        cell.value=val; cell.font=Font(name="Arial",bold=True,size=10,color="FFFFFFFF")
        cell.fill=PatternFill("solid",fgColor=CG); cell.alignment=Alignment(horizontal="right",vertical="center"); cell.border=bd(); cell.number_format=FMT
    def judul(ws,title,period,nc):
        lc=get_column_letter(nc); ws.merge_cells(f"A1:{lc}1"); ws["A1"].value=title
        ws["A1"].font=Font(name="Arial",bold=True,size=16,color=CH); ws["A1"].alignment=Alignment(horizontal="center",vertical="center"); ws.row_dimensions[1].height=36
        ws.merge_cells(f"A2:{lc}2"); ws["A2"].value=f"Periode: {period}  |  Diekspor: {datetime.now().strftime('%d %B %Y %H:%M')}"
        ws["A2"].font=Font(name="Arial",size=9,italic=True,color="FF928E89"); ws["A2"].alignment=Alignment(horizontal="center"); ws.row_dimensions[2].height=18; ws.row_dimensions[3].height=6

    wst={"selesai":CG,"dibatalkan":CR,"dikirim":"FFE76F00","diproses":"FF1A4A8A","pending":"FF928E89"}
    per = bulan if bulan else "Semua Data"

    # Sheet 1
    ws1=wb.active; ws1.title="Laporan Penjualan"
    judul(ws1,"LAPORAN PENJUALAN",per,10)
    for c,h in enumerate(["No","No. Pesanan","Pelanggan","Kota","Tgl. Pesan","Tgl. Kirim","Status","Total Harga (Rp)","Diskon (%)","Total Bayar (Rp)"],1): sh(ws1.cell(4,c),h)
    ws1.row_dimensions[4].height=26
    for i,p in enumerate(dp):
        r=i+5; bg=None if i%2==0 else CST
        tp=p["tanggal_pesan"].strftime("%d/%m/%Y") if p["tanggal_pesan"] else "-"
        tk=p["tanggal_kirim"].strftime("%d/%m/%Y") if p["tanggal_kirim"] else "-"
        sd(ws1.cell(r,1),i+1,"center",bg=bg); sd(ws1.cell(r,2),p["no_pesanan"],"center",True,bg); sd(ws1.cell(r,3),p["pelanggan"],"left",bg=bg); sd(ws1.cell(r,4),p["kota"] or "-","left",bg=bg)
        sd(ws1.cell(r,5),tp,"center",bg=bg); sd(ws1.cell(r,6),tk,"center",bg=bg); sd(ws1.cell(r,7),p["status"].upper(),"center",True,bg,wst.get(p["status"],CH))
        sd(ws1.cell(r,8),float(p["total_harga"]),"right",bg=bg); sd(ws1.cell(r,9),float(p["diskon"]),"right",bg=bg); sd(ws1.cell(r,10),float(p["total_bayar"]),"right",True,bg,CG); ws1.row_dimensions[r].height=18
    rt=len(dp)+5; lr=rt-1
    ws1.merge_cells(f"A{rt}:G{rt}"); cc=ws1[f"A{rt}"]; cc.value=f"TOTAL ({len(dp)} TRANSAKSI)"; cc.font=Font(name="Arial",bold=True,size=10,color="FFFFFFFF"); cc.fill=PatternFill("solid",fgColor=CH); cc.alignment=Alignment(horizontal="right",vertical="center"); cc.border=bd()
    st(ws1.cell(rt,8),f"=SUM(H5:H{lr})"); st(ws1.cell(rt,9),f"=AVERAGE(I5:I{lr})"); st(ws1.cell(rt,10),f"=SUM(J5:J{lr})"); ws1.row_dimensions[rt].height=24
    for row in ws1.iter_rows(5,rt,8,10):
        for cell in row: cell.number_format=FMT
    for c,w in enumerate([5,15,28,15,12,12,12,18,11,18],1): ws1.column_dimensions[get_column_letter(c)].width=w
    ws1.freeze_panes="A5"

    # Sheet 2
    ws2=wb.create_sheet("Keluar Masuk Stok"); judul(ws2,"LAPORAN KELUAR MASUK STOK",per,10)
    for c,h in enumerate(["No","Tanggal","No. Pesanan","Pelanggan","Kode Produk","Nama Produk","Satuan","Qty Keluar","Harga Satuan (Rp)","Subtotal (Rp)"],1): sh(ws2.cell(4,c),h)
    ws2.row_dimensions[4].height=26
    for i,item in enumerate(ds):
        r=i+5; bg=None if i%2==0 else CST
        tg=item["tanggal_pesan"].strftime("%d/%m/%Y") if item["tanggal_pesan"] else "-"
        sd(ws2.cell(r,1),i+1,"center",bg=bg); sd(ws2.cell(r,2),tg,"center",bg=bg); sd(ws2.cell(r,3),item["no_pesanan"],"center",True,bg); sd(ws2.cell(r,4),item["pelanggan"],"left",bg=bg)
        sd(ws2.cell(r,5),item["kode_produk"],"center",bg=bg); sd(ws2.cell(r,6),item["produk"],"left",bg=bg); sd(ws2.cell(r,7),item["satuan"],"center",bg=bg)
        sd(ws2.cell(r,8),item["jumlah"],"right",True,bg,CR); sd(ws2.cell(r,9),float(item["harga_satuan"]),"right",bg=bg); sd(ws2.cell(r,10),float(item["subtotal"]),"right",True,bg); ws2.row_dimensions[r].height=18
    rt2=len(ds)+5; lr2=rt2-1
    ws2.merge_cells(f"A{rt2}:G{rt2}"); cc2=ws2[f"A{rt2}"]; cc2.value=f"TOTAL ({len(ds)} ITEM)"; cc2.font=Font(name="Arial",bold=True,size=10,color="FFFFFFFF"); cc2.fill=PatternFill("solid",fgColor=CH); cc2.alignment=Alignment(horizontal="right",vertical="center"); cc2.border=bd()
    st(ws2.cell(rt2,8),f"=SUM(H5:H{lr2})"); st(ws2.cell(rt2,9),f"=AVERAGE(I5:I{lr2})"); st(ws2.cell(rt2,10),f"=SUM(J5:J{lr2})"); ws2.row_dimensions[rt2].height=24
    for row in ws2.iter_rows(5,rt2,9,10):
        for cell in row: cell.number_format=FMT
    for c,w in enumerate([5,12,15,28,14,28,9,11,18,16],1): ws2.column_dimensions[get_column_letter(c)].width=w
    ws2.freeze_panes="A5"

    # Sheet 3
    ws3=wb.create_sheet("Ringkasan Stok"); judul(ws3,"RINGKASAN STOK PRODUK",f"Per {datetime.now().strftime('%d %B %Y')}",9)
    for c,h in enumerate(["No","Kode","Nama Produk","Kategori","Satuan","Harga (Rp)","Stok","Min. Stok","Nilai Stok (Rp)"],1): sh(ws3.cell(4,c),h)
    ws3.row_dimensions[4].height=26
    for i,p in enumerate(dr):
        r=i+5; kritis=p["stok"]<=p["stok_minimum"]; bg="FFFFF3CD" if kritis else (None if i%2==0 else CST)
        sd(ws3.cell(r,1),i+1,"center",bg=bg); sd(ws3.cell(r,2),p["kode_produk"],"center",True,bg); sd(ws3.cell(r,3),p["nama"],"left",bg=bg); sd(ws3.cell(r,4),p["kategori"] or "-","left",bg=bg)
        sd(ws3.cell(r,5),p["satuan"],"center",bg=bg); sd(ws3.cell(r,6),float(p["harga"]),"right",bg=bg); sd(ws3.cell(r,7),p["stok"],"right",True,bg,CR if kritis else CG); sd(ws3.cell(r,8),p["stok_minimum"],"right",bg=bg); sd(ws3.cell(r,9),float(p["nilai_stok"]),"right",True,bg); ws3.row_dimensions[r].height=18
    rt3=len(dr)+5; lr3=rt3-1
    ws3.merge_cells(f"A{rt3}:H{rt3}"); cc3=ws3[f"A{rt3}"]; cc3.value=f"TOTAL NILAI STOK ({len(dr)} PRODUK)"; cc3.font=Font(name="Arial",bold=True,size=10,color="FFFFFFFF"); cc3.fill=PatternFill("solid",fgColor=CH); cc3.alignment=Alignment(horizontal="right",vertical="center"); cc3.border=bd()
    st(ws3.cell(rt3,9),f"=SUM(I5:I{lr3})"); ws3.row_dimensions[rt3].height=24
    ket=rt3+2; ws3.merge_cells(f"A{ket}:F{ket}"); k=ws3[f"A{ket}"]; k.value="Keterangan: Baris KUNING = stok di bawah minimum, perlu restok!"; k.font=Font(name="Arial",size=8,italic=True,color="FF7A4500"); k.fill=PatternFill("solid",fgColor="FFFFF3CD")
    for row in ws3.iter_rows(5,rt3,6,9):
        for cell in row:
            if cell.column in [6,9]: cell.number_format=FMT
    for c,w in enumerate([5,14,30,16,9,16,9,11,20],1): ws3.column_dimensions[get_column_letter(c)].width=w
    ws3.freeze_panes="A5"

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    nama=f"Laporan_PendidikanStore_{bulan if bulan else 'SemuaData'}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=nama)


# =============================================================================
# [PYTHON] -- BAGIAN 15: NOTIFIKASI EMAIL STOK
# =============================================================================
@app.route("/notifikasi/stok")
def notifikasi_stok():
    cur = mysql.connection.cursor()
    cur.execute("SELECT nama, stok, stok_minimum, satuan FROM produk WHERE stok<=stok_minimum AND status='aktif' ORDER BY stok ASC")
    pm = cur.fetchall()
    cur.close()
    if pm:
        kirim_email_stok(pm)
        flash(f"Email terkirim! {len(pm)} produk stok menipis.", "warning")
    else:
        flash("Semua stok aman!", "success")
    return redirect("/produk")



# =============================================================================
# [SQL] -- TABEL USERS (jalankan di phpMyAdmin dulu!)
# =============================================================================
# ALTER DATABASE sistem_penjualan CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci;
#
# CREATE TABLE IF NOT EXISTS users (
#     id           INT AUTO_INCREMENT PRIMARY KEY,
#     nama         VARCHAR(100) NOT NULL,
#     email        VARCHAR(100) UNIQUE NOT NULL,
#     password     VARCHAR(255) NOT NULL,       -- password terenkripsi (hash)
#     role         ENUM("admin","staff") DEFAULT "staff",
#     avatar       VARCHAR(10) DEFAULT "A",     -- inisial untuk avatar
#     provider     VARCHAR(20) DEFAULT "email", -- "email" atau "google"
#     status       ENUM("aktif","nonaktif") DEFAULT "aktif",
#     last_login   TIMESTAMP NULL,
#     created_at   TIMESTAMP DEFAULT CURRENT_TIMESTAMP
# );
#
# -- Akun admin default (password: admin123)
# INSERT INTO users (nama, email, password, role, avatar) VALUES
# ("Administrator", "admin@pendidikanstore.com",
#  "pbkdf2:sha256:600000$saltkey$hashedpassword", "admin", "A");
# -- PENTING: Jalankan route /setup-admin untuk buat akun admin dengan hash yang benar!
# =============================================================================


# =============================================================================
# [PYTHON] -- LOGIN SYSTEM: Decorator & Helper
# =============================================================================

# [CSS] Styling khusus halaman login ΓÇö terpisah dari CSS_GLOBAL
CSS_LOGIN = """
/* [CSS] Halaman login: full screen split layout */
* { box-sizing: border-box; margin: 0; padding: 0; }

body {
    font-family: "IBM Plex Sans", sans-serif;
    background: #052e16;
    min-height: 100vh;
    display: flex;
    align-items: center;
    justify-content: center;
}

/* [CSS] Container utama: dua kolom */
.login-wrapper {
    display: flex;
    width: 920px;
    min-height: 560px;
    border-radius: 16px;
    overflow: hidden;
    box-shadow: 0 32px 80px rgba(0,0,0,0.5);
}

/* [CSS] Panel kiri ΓÇö branding & ilustrasi */
.login-left {
    flex: 1;
    background: #14532d;
    padding: 48px 40px;
    display: flex;
    flex-direction: column;
    justify-content: space-between;
    position: relative;
    overflow: hidden;
}

/* [CSS] Dekorasi background geometris */
.login-left::before {
    content: "";
    position: absolute;
    top: -80px; right: -80px;
    width: 320px; height: 320px;
    border-radius: 50%;
    background: rgba(74,222,128,0.08);
}
.login-left::after {
    content: "";
    position: absolute;
    bottom: -60px; left: -60px;
    width: 240px; height: 240px;
    border-radius: 50%;
    background: rgba(74,222,128,0.05);
}

/* [CSS] Logo aplikasi */
.login-logo {
    display: flex;
    align-items: center;
    gap: 12px;
    position: relative;
    z-index: 1;
}
.login-logo-icon {
    width: 44px; height: 44px;
    background: #4ade80;
    border-radius: 10px;
    display: flex;
    align-items: center;
    justify-content: center;
    font-size: 22px;
    color: #052e16;
    font-weight: 900;
}
.login-logo-text { color: #fff; }
.login-logo-name { font-size: 18px; font-weight: 700; letter-spacing: -0.02em; }
.login-logo-sub  { font-size: 11px; color: rgba(255,255,255,0.4); letter-spacing: 0.08em; text-transform: uppercase; }

/* [CSS] Teks promo di panel kiri */
.login-hero { position: relative; z-index: 1; }
.login-hero h2 {
    font-size: 28px; font-weight: 800;
    color: #fff; line-height: 1.3;
    letter-spacing: -0.03em;
    margin-bottom: 12px;
}
.login-hero h2 span { color: #4ade80; }
.login-hero p { font-size: 13px; color: rgba(255,255,255,0.45); line-height: 1.7; }

/* [CSS] Fitur list di bawah */
.login-features { position: relative; z-index: 1; }
.login-feature {
    display: flex;
    align-items: center;
    gap: 10px;
    margin-bottom: 10px;
    font-size: 12px;
    color: rgba(255,255,255,0.5);
}
.login-feature-dot {
    width: 6px; height: 6px;
    border-radius: 50%;
    background: #4ade80;
    flex-shrink: 0;
}

/* [CSS] Panel kanan ΓÇö form login */
.login-right {
    width: 400px;
    background: #f0fdf4;
    padding: 48px 40px;
    display: flex;
    flex-direction: column;
    justify-content: center;
}

.login-title    { font-size: 22px; font-weight: 800; color: #052e16; margin-bottom: 4px; letter-spacing: -0.02em; }
.login-subtitle { font-size: 13px; color: #4ade80; margin-bottom: 32px; }

/* [CSS] Tombol Login dengan Google */
.btn-google {
    display: flex; align-items: center; justify-content: center;
    gap: 10px;
    width: 100%; padding: 11px;
    background: #fff;
    border: 1.5px solid #bbf7d0;
    border-radius: 8px;
    font-size: 13px; font-weight: 600;
    color: #052e16; cursor: pointer;
    text-decoration: none;
    transition: all 0.15s;
    margin-bottom: 20px;
}
.btn-google:hover { background: #f9f8f6; border-color: #86efac; box-shadow: 0 2px 8px rgba(0,0,0,0.08); }
.btn-google svg  { flex-shrink: 0; }

/* [CSS] Divider "atau" */
.login-divider {
    display: flex; align-items: center;
    gap: 12px; margin-bottom: 20px;
}
.login-divider-line { flex: 1; height: 1px; background: #bbf7d0; }
.login-divider-text { font-size: 11px; color: #4ade80; font-weight: 600; letter-spacing: 0.06em; text-transform: uppercase; }

/* [CSS] Form group */
.form-group-login { margin-bottom: 16px; }
.form-group-login label {
    display: block; font-size: 11px; font-weight: 700;
    letter-spacing: 0.08em; text-transform: uppercase;
    color: #166534; margin-bottom: 6px;
}
.form-group-login input {
    width: 100%; padding: 11px 14px;
    border: 1.5px solid #bbf7d0;
    border-radius: 8px;
    font-size: 13px; color: #052e16;
    background: #fff;
    font-family: "IBM Plex Sans", sans-serif;
    outline: none; transition: all 0.15s;
}
.form-group-login input:focus {
    border-color: #052e16;
    box-shadow: 0 0 0 3px rgba(26,25,22,0.08);
}

/* [CSS] Row ingat saya + lupa password */
.login-options {
    display: flex; align-items: center;
    justify-content: space-between;
    margin-bottom: 20px; font-size: 12px;
}
.login-options label { display: flex; align-items: center; gap: 6px; color: #166534; cursor: pointer; font-size: 12px; letter-spacing: 0; text-transform: none; }
.login-options a     { color: #052e16; font-weight: 600; text-decoration: none; }
.login-options a:hover { text-decoration: underline; }

/* [CSS] Tombol submit login */
.btn-login {
    width: 100%; padding: 12px;
    background: #14532d; color: #fff;
    border: none; border-radius: 8px;
    font-size: 14px; font-weight: 700;
    cursor: pointer; font-family: "IBM Plex Sans", sans-serif;
    transition: all 0.15s; letter-spacing: 0.01em;
}
.btn-login:hover { background: #15803d; transform: translateY(-1px); box-shadow: 0 4px 12px rgba(0,0,0,0.2); }
.btn-login:active { transform: translateY(0); }

/* [CSS] Link daftar akun baru */
.login-register { text-align: center; margin-top: 20px; font-size: 12px; color: #4ade80; }
.login-register a { color: #052e16; font-weight: 700; text-decoration: none; }
.login-register a:hover { text-decoration: underline; }

/* [CSS] Flash message di halaman login */
.login-flash {
    padding: 10px 14px; border-radius: 8px;
    margin-bottom: 16px; font-size: 12px; font-weight: 600;
}
.login-flash.danger  { background: #fde8e8; color: #c0392b; border-left: 3px solid #c0392b; }
.login-flash.warning { background: #fff3cd; color: #e76f00; border-left: 3px solid #e76f00; }
.login-flash.success { background: #d8f3dc; color: #2d6a4f; border-left: 3px solid #2d6a4f; }

/* [CSS] Animasi fade in saat halaman dimuat */
@keyframes fadeUp {
    from { opacity: 0; transform: translateY(20px); }
    to   { opacity: 1; transform: translateY(0); }
}
.login-right { animation: fadeUp 0.5s ease forwards; }

/* [CSS] Responsive mobile */
@media (max-width: 768px) {
    .login-wrapper { flex-direction: column; width: 95%; min-height: auto; }
    .login-left    { padding: 32px 28px; min-height: auto; }
    .login-right   { width: 100%; padding: 32px 28px; }
    .login-hero h2 { font-size: 22px; }
    .login-features { display: none; }
}
"""

# [HTML] Template halaman login
LOGIN_TEMPLATE = """
<!DOCTYPE html>
<html lang="id">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Login ΓÇö PendidikanStore APSI</title>
    <link href="https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@400;500;600;700;800&display=swap" rel="stylesheet">
    <style>{{ css|safe }}</style>
</head>
<body>

<div class="login-wrapper">

    <!-- [HTML] Panel Kiri: Branding -->
    <div class="login-left">
        <!-- Logo -->
        <div class="login-logo">
            <div class="login-logo-icon">S</div>
            <div class="login-logo-text">
                <div class="login-logo-name">PendidikanStore</div>
                <div class="login-logo-sub">Perlengkapan Siswa & Mahasiswa</div>
            </div>
        </div>

        <!-- Hero text -->
        <div class="login-hero">
            <h2>Kelola Toko Pendidikan<br>dengan <span>Lebih Cerdas</span></h2>
            <p>Toko alat tulis dan perlengkapan pendidikan lengkap untuk TK, SD, SMP, SMA/SMK, hingga mahasiswa.</p>
        </div>

        <!-- Fitur list -->
        <div class="login-features">
            <div class="login-feature">
                <div class="login-feature-dot"></div>
                Pensil, pena, buku & alat tulis lengkap
            </div>
            <div class="login-feature">
                <div class="login-feature-dot"></div>
                Produk untuk semua jenjang pendidikan
            </div>
            <div class="login-feature">
                <div class="login-feature-dot"></div>
                Laporan penjualan & stok otomatis
            </div>
            <div class="login-feature">
                <div class="login-feature-dot"></div>
                Notifikasi stok & pesanan real-time
            </div>
        </div>
    </div>

    <!-- [HTML] Panel Kanan: Form Login -->
    <div class="login-right">
        <h1 class="login-title">Selamat datang</h1>
        <p class="login-subtitle">Masuk ke akun PendidikanStore kamu</p>

        <!-- [HTML] Flash messages -->
        {% with messages = get_flashed_messages(with_categories=true) %}
            {% if messages %}
                {% for category, message in messages %}
                    <div class="login-flash {{ category }}">{{ message }}</div>
                {% endfor %}
            {% endif %}
        {% endwith %}

        <!-- [HTML] Tombol Login dengan Google -->
        <a href="/login/google" class="btn-google">
            <!-- [HTML] Logo Google SVG -->
            <svg width="18" height="18" viewBox="0 0 48 48">
                <path fill="#EA4335" d="M24 9.5c3.54 0 6.71 1.22 9.21 3.6l6.85-6.85C35.9 2.38 30.47 0 24 0 14.62 0 6.51 5.38 2.56 13.22l7.98 6.19C12.43 13.72 17.74 9.5 24 9.5z"/>
                <path fill="#4285F4" d="M46.98 24.55c0-1.57-.15-3.09-.38-4.55H24v9.02h12.94c-.58 2.96-2.26 5.48-4.78 7.18l7.73 6c4.51-4.18 7.09-10.36 7.09-17.65z"/>
                <path fill="#FBBC05" d="M10.53 28.59c-.48-1.45-.76-2.99-.76-4.59s.27-3.14.76-4.59l-7.98-6.19C.92 16.46 0 20.12 0 24c0 3.88.92 7.54 2.56 10.78l7.97-6.19z"/>
                <path fill="#34A853" d="M24 48c6.48 0 11.93-2.13 15.89-5.81l-7.73-6c-2.18 1.48-4.97 2.35-8.16 2.35-6.26 0-11.57-4.22-13.47-9.91l-7.98 6.19C6.51 42.62 14.62 48 24 48z"/>
            </svg>
            Lanjutkan dengan Google
        </a>

        <!-- [HTML] Divider -->
        <div class="login-divider">
            <div class="login-divider-line"></div>
            <span class="login-divider-text">atau</span>
            <div class="login-divider-line"></div>
        </div>

        <!-- [HTML] Form email & password -->
        <form method="POST" action="/login">
            <div class="form-group-login">
                <label>Email</label>
                <input type="email" name="email" placeholder="nama@perusahaan.com"
                       value="{{ email or "" }}" required autocomplete="email">
            </div>
            <div class="form-group-login">
                <label>Password</label>
                <input type="password" name="password" placeholder="Masukkan password"
                       required autocomplete="current-password">
            </div>

            <!-- [HTML] Opsi ingat saya & lupa password -->
            <div class="login-options">
                <label>
                    <input type="checkbox" name="remember"> Ingat saya
                </label>
                <a href="/lupa-password">Lupa password?</a>
            </div>

            <button type="submit" class="btn-login">Masuk ke PendidikanStore</button>
        </form>

        <!-- [HTML] Link registrasi -->
        <div class="login-register">
            Belum punya akun? <a href="/register">Daftar sekarang</a>
        </div>
    </div>
</div>

<!-- [JAVASCRIPT] Auto-hide flash setelah 4 detik -->
<script>
document.querySelectorAll(".login-flash").forEach(function(el) {
    setTimeout(function() {
        el.style.opacity = "0";
        el.style.transition = "opacity 0.4s";
        setTimeout(function() { el.remove(); }, 400);
    }, 4000);
});
</script>
</body>
</html>
"""

# [HTML] Template halaman registrasi
REGISTER_TEMPLATE = """
<!DOCTYPE html>
<html lang="id">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Daftar Akun ΓÇö PendidikanStore</title>
    <link href="https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@400;500;600;700;800&display=swap" rel="stylesheet">
    <style>{{ css|safe }}</style>
</head>
<body>
<div class="login-wrapper">
    <div class="login-left">
        <div class="login-logo">
            <div class="login-logo-icon">S</div>
            <div class="login-logo-text">
                <div class="login-logo-name">PendidikanStore</div>
                <div class="login-logo-sub">Perlengkapan Siswa & Mahasiswa</div>
            </div>
        </div>
        <div class="login-hero">
            <h2>Buat Akun <span>Baru</span></h2>
            <p>Daftarkan diri kamu untuk mengakses sistem penjualan PendidikanStore.</p>
        </div>
        <div class="login-features">
            <div class="login-feature"><div class="login-feature-dot"></div>Akses semua fitur sistem</div>
            <div class="login-feature"><div class="login-feature-dot"></div>Data tersimpan aman & terenkripsi</div>
            <div class="login-feature"><div class="login-feature-dot"></div>Multi user dengan role berbeda</div>
        </div>
    </div>
    <div class="login-right">
        <h1 class="login-title">Buat Akun</h1>
        <p class="login-subtitle">Isi data di bawah untuk mendaftar</p>

        {% with messages = get_flashed_messages(with_categories=true) %}
            {% if messages %}
                {% for category, message in messages %}
                    <div class="login-flash {{ category }}">{{ message }}</div>
                {% endfor %}
            {% endif %}
        {% endwith %}

        <form method="POST" action="/register">
            <div class="form-group-login">
                <label>Nama Lengkap</label>
                <input type="text" name="nama" placeholder="John Doe" required>
            </div>
            <div class="form-group-login">
                <label>Email</label>
                <input type="email" name="email" placeholder="nama@perusahaan.com" required>
            </div>
            <div class="form-group-login">
                <label>Password</label>
                <input type="password" name="password" placeholder="Minimal 6 karakter" required minlength="6">
            </div>
            <div class="form-group-login">
                <label>Konfirmasi Password</label>
                <input type="password" name="confirm_password" placeholder="Ulangi password" required>
            </div>
            <button type="submit" class="btn-login" style="margin-top:4px">Buat Akun</button>
        </form>
        <div class="login-register">
            Sudah punya akun? <a href="/login">Masuk di sini</a>
        </div>
    </div>
</div>
</body>
</html>
"""


# =============================================================================
# [PYTHON] -- ROUTE: SETUP ADMIN (jalankan sekali untuk buat akun admin)
# Buka: http://localhost:5000/setup-admin
# =============================================================================
@app.route("/setup-admin")


@app.route('/api/cek-stok')
def api_cek_stok():
    """
    [PYTHON] API endpoint untuk polling notifikasi real-time.
    Dipanggil JavaScript setiap 30 detik untuk cek stok menipis & pesanan baru.
    """
    cur = mysql.connection.cursor()

    # [SQL] Cek produk stok menipis
    cur.execute("""
        SELECT nama, stok, satuan FROM produk
        WHERE stok <= stok_minimum AND status='aktif'
        ORDER BY stok ASC LIMIT 5
    """)
    stok_menipis = cur.fetchall()

    # [SQL] Hitung pesanan baru hari ini
    cur.execute("""
        SELECT COUNT(*) as total FROM pesanan
        WHERE DATE(created_at) = CURDATE() AND status='pending'
    """)
    pesanan_baru = cur.fetchone()['total']

    cur.close()

    return jsonify({
        'stok_menipis': list(stok_menipis),
        'pesanan_baru': pesanan_baru
    })


# =============================================================================
# [PYTHON] -- FITUR BARU 2: UPLOAD GAMBAR PRODUK
# =============================================================================

import os
import uuid
# os   -> untuk operasi file system (buat folder, cek file)
# uuid -> generate nama file unik agar tidak bertabrakan

# [PYTHON] Folder penyimpanan gambar produk
UPLOAD_FOLDER = 'static/uploads/produk'


# Route untuk serve gambar produk
from flask import send_from_directory

@app.route('/static/uploads/produk/<filename>')
def uploaded_file(filename):
    """[PYTHON] Serve file gambar yang sudah diupload."""
    return send_from_directory(UPLOAD_FOLDER, filename)


# =============================================================================
# [PYTHON] -- FITUR BARU 3: KATALOG PRODUK (Marketplace Style)
# =============================================================================

@app.route('/katalog')
def katalog():
    """
    [PYTHON] Halaman katalog produk bergaya marketplace (Shopee/Tokopedia).
    Menampilkan gambar produk, harga, dan stok dalam grid card.
    """
    cur = mysql.connection.cursor()

    kategori_filter = request.args.get('kategori', '')
    search          = request.args.get('search', '')
    sort            = request.args.get('sort', 'nama')

    # [SQL] Query produk dengan filter dan sorting
    query  = """
        SELECT p.*, k.nama as nama_kategori
        FROM produk p LEFT JOIN kategori k ON p.kategori_id = k.id
        WHERE p.status = 'aktif'
    """
    params = []

    if kategori_filter:
        query += " AND p.kategori_id = %s"
        params.append(kategori_filter)

    if search:
        query += " AND p.nama LIKE %s"
        params.append(f'%{search}%')

    # Sorting options
    sort_map = {
        'nama':        'p.nama ASC',
        'harga_asc':   'p.harga ASC',
        'harga_desc':  'p.harga DESC',
        'stok':        'p.stok DESC',
    }
    query += f" ORDER BY {sort_map.get(sort, 'p.nama ASC')}"

    cur.execute(query, params)
    produk_list = cur.fetchall()

    # [SQL] Ambil semua kategori untuk filter
    cur.execute("SELECT * FROM kategori ORDER BY nama")
    kategori_list = cur.fetchall()

    # [SQL] Statistik singkat
    cur.execute("SELECT COUNT(*) as total FROM produk WHERE status='aktif'")
    total_produk = cur.fetchone()['total']

    cur.close()

    # [PYTHON] Bangun kartu produk
    def fmt_harga(h):
        return f"Rp {int(h):,}".replace(',', '.')

    # [PYTHON] Build filter kategori buttons
    filter_html = f'''
    <a href="/katalog?sort={sort}" class="filter-btn {"active" if not kategori_filter else ""}">
        Semua ({total_produk})
    </a>'''
    for k in kategori_list:
        active = 'active' if str(k['id']) == str(kategori_filter) else ''
        filter_html += f'''
        <a href="/katalog?kategori={k["id"]}&sort={sort}"
           class="filter-btn {active}">{k["nama"]}</a>'''

    # [PYTHON] Sort options
    sort_options = {
        'nama': 'Nama A-Z',
        'harga_asc': 'Harga Termurah',
        'harga_desc': 'Harga Termahal',
        'stok': 'Stok Terbanyak'
    }
    sort_html = '''<select onchange="window.location='/katalog?sort='+this.value+'&kategori=' + ''' + kategori_filter + '''"
        style="padding:6px 10px;border:1.5px solid var(--border);border-radius:6px;
               font-size:12px;font-family:inherit;background:var(--surface);color:var(--text);outline:none">'''
    for val, label in sort_options.items():
        selected = 'selected' if sort == val else ''
        sort_html += f'<option value="{val}" {selected}>{label}</option>'
    sort_html += '</select>'

    # [PYTHON] Build product cards
    cards_html = ''
    for p in produk_list:
        harga      = fmt_harga(p['harga'])
        stok_kritis = p['stok'] <= p['stok_minimum']
        badge_html = ''
        if stok_kritis:
            badge_html = '<span class="katalog-badge">Stok Tipis</span>'
        elif p['stok'] > 50:
            badge_html = '<span class="katalog-badge green">Ready</span>'

        # Gambar produk
        if p.get('gambar'):
            img_html = f'<img src="/static/uploads/produk/{p["gambar"]}" alt="{p["nama"]}" loading="lazy">'
        else:
            # Placeholder emoji berdasarkan kategori
            emoji_map = {'Elektronik': '≡ƒÆ╗', 'Pakaian': '≡ƒæò', 'Makanan': '≡ƒì£', 'Alat Tulis': 'Γ£Å∩╕Å'}
            emoji = emoji_map.get(p['nama_kategori'], '≡ƒôª')
            img_html = f'<div class="katalog-img-placeholder">{emoji}</div>'

        warna_harga = 'color:var(--red)' if stok_kritis else 'color:var(--red)'

        cards_html += f'''
        <a href="/produk/edit/{p["id"]}" class="katalog-card">
            <div class="katalog-img">
                {badge_html}
                {img_html}
            </div>
            <div class="katalog-info">
                <div class="katalog-kategori">{p["nama_kategori"] or "Umum"}</div>
                <div class="katalog-nama">{p["nama"]}</div>
                <div class="katalog-harga">{harga}</div>
                <div class="katalog-stok">Stok: {p["stok"]} {p["satuan"]}</div>
            </div>
            <div class="katalog-footer">
                <span>{p["kode_produk"]}</span>
                <span style="font-size:11px;color:{"var(--red)" if stok_kritis else "var(--green)"}">
                    {"ΓÜá Menipis" if stok_kritis else "Γ£ô Tersedia"}
                </span>
            </div>
        </a>'''

    if not cards_html:
        cards_html = '''
        <div style="grid-column:1/-1;text-align:center;padding:60px;color:var(--text-3)">
            <div style="font-size:48px;margin-bottom:12px">≡ƒöì</div>
            <p style="font-size:14px">Tidak ada produk ditemukan</p>
        </div>'''

    content = f'''
    <div class="page-header">
        <div>
            <h2>Katalog Produk</h2>
            <div class="breadcrumb">Produk / <span>Katalog</span></div>
        </div>
        <div style="display:flex;gap:8px;align-items:center">
            <!-- [HTML] Search produk -->
            <form method="GET" action="/katalog" style="display:flex;gap:8px">
                <input type="hidden" name="kategori" value="{kategori_filter}">
                <input type="hidden" name="sort" value="{sort}">
                <input class="search-input" type="text" name="search"
                       placeholder="Cari produk..." value="{search}" style="min-width:200px">
                <button type="submit" class="btn btn-outline btn-sm">Cari</button>
            </form>
            {sort_html}
            <a href="/produk/tambah" class="btn btn-primary btn-sm">+ Tambah Produk</a>
        </div>
    </div>

    <!-- [HTML] Filter kategori -->
    <div class="katalog-filter">
        {filter_html}
    </div>

    <!-- [HTML] Info jumlah produk -->
    <div style="font-size:12px;color:var(--text-3);margin-bottom:12px">
        Menampilkan <strong style="color:var(--text)">{len(produk_list)}</strong> produk
        {f"dalam kategori <strong>{kategori_filter}</strong>" if kategori_filter else ""}
    </div>

    <!-- [HTML] Grid kartu produk marketplace -->
    <div class="katalog-grid">
        {cards_html}
    </div>
    '''

    return render_page('Katalog Produk', 'produk', content)


# [PYTHON] Update route tambah/edit produk untuk support upload gambar
@app.route('/produk/tambah-foto', methods=['GET', 'POST'])
def produk_tambah_foto():
    """
    [PYTHON] Form tambah produk dengan fitur upload foto.
    Menggunakan enctype multipart/form-data untuk upload file.
    """
    cur = mysql.connection.cursor()

    if request.method == 'POST':
        kode         = generate_kode('PRD', 'produk', 'kode_produk')
        nama         = request.form['nama']
        kategori_id  = request.form['kategori_id'] or None
        harga        = request.form['harga']
        stok         = request.form['stok']
        stok_minimum = request.form['stok_minimum']
        satuan       = request.form['satuan']
        deskripsi    = request.form.get('deskripsi', '')

        # [PYTHON] Proses upload gambar jika ada
        gambar = None
        if 'gambar' in request.files:
            file = request.files['gambar']
            if file.filename:
                gambar = save_gambar(file)

        # [SQL] Cek apakah kolom gambar sudah ada
        try:
            cur.execute("""
                ALTER TABLE produk ADD COLUMN IF NOT EXISTS gambar VARCHAR(255) NULL
            """)
            mysql.connection.commit()
        except:
            pass

        cur.execute("""
            INSERT INTO produk (kode_produk, nama, kategori_id, harga, stok,
                               stok_minimum, satuan, deskripsi, gambar)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (kode, nama, kategori_id, harga, stok, stok_minimum, satuan, deskripsi, gambar))
        mysql.connection.commit()
        cur.close()

        # [JAVASCRIPT] Kirim notifikasi produk baru
        flash(f'Produk {nama} berhasil ditambahkan!', 'success')
        return redirect('/katalog')

    cur.execute("SELECT * FROM kategori")
    kategori = cur.fetchall()
    cur.close()

    opt_kat = '<option value="">ΓÇö Pilih Kategori ΓÇö</option>'
    for k in kategori:
        opt_kat += f'<option value="{k["id"]}">{k["nama"]}</option>'

    content = f'''
    <div class="page-header">
        <div><h2>Tambah Produk + Foto</h2>
        <div class="breadcrumb"><a href="/katalog">Katalog</a> / Tambah</div></div>
    </div>
    <div class="form-card" style="max-width:700px">
        <div class="form-section-title">Informasi Produk & Foto</div>
        <!-- [HTML] enctype multipart/form-data wajib untuk upload file -->
        <form method="POST" enctype="multipart/form-data">
            <div class="form-grid">
                <!-- Upload foto -->
                <div class="form-group full">
                    <label>Foto Produk</label>
                    <div class="upload-area" onclick="document.getElementById('inputGambar').click()">
                        <div id="previewContainer">
                            <div style="font-size:32px">≡ƒô╖</div>
                            <p style="font-size:12px;color:var(--text-3);margin-top:8px">
                                Klik untuk upload foto produk<br>
                                <span style="font-size:10px">PNG, JPG, WEBP ΓÇö Max 5MB</span>
                            </p>
                        </div>
                        <!-- [HTML] Input file tersembunyi -->
                        <input type="file" id="inputGambar" name="gambar"
                               accept="image/*" style="display:none"
                               onchange="previewGambar(this)">
                    </div>
                </div>
                <div class="form-group full">
                    <label>Nama Produk *</label>
                    <input type="text" name="nama" required placeholder="Nama lengkap produk">
                </div>
                <div class="form-group">
                    <label>Kategori</label>
                    <select name="kategori_id">{opt_kat}</select>
                </div>
                <div class="form-group">
                    <label>Satuan</label>
                    <select name="satuan">
                        <option value="pcs">pcs</option>
                        <option value="unit">unit</option>
                        <option value="kg">kg</option>
                        <option value="liter">liter</option>
                        <option value="box">box</option>
                    </select>
                </div>
                <div class="form-group">
                    <label>Harga Jual (Rp) *</label>
                    <input type="number" name="harga" required placeholder="0" min="0">
                </div>
                <div class="form-group">
                    <label>Stok Awal</label>
                    <input type="number" name="stok" value="0" min="0">
                </div>
                <div class="form-group">
                    <label>Stok Minimum</label>
                    <input type="number" name="stok_minimum" value="5" min="0">
                </div>
                <div class="form-group full">
                    <label>Deskripsi Produk</label>
                    <textarea name="deskripsi" placeholder="Deskripsikan produk kamu..."></textarea>
                </div>
            </div>
            <div class="form-actions">
                <button type="submit" class="btn btn-primary">Simpan Produk</button>
                <a href="/katalog" class="btn btn-outline">Batal</a>
            </div>
        </form>
    </div>
    '''

    # [JAVASCRIPT] Preview gambar sebelum upload
    extra_scripts = '''
    <script>
    function previewGambar(input) {
        if (input.files && input.files[0]) {
            var reader = new FileReader();
            reader.onload = function(e) {
                var preview = document.getElementById("previewContainer");
                preview.innerHTML =
                    "<img src='" + e.target.result + "' class='upload-preview'>" +
                    "<p style='font-size:11px;color:var(--text-3);margin-top:8px'>" +
                    input.files[0].name + "</p>";
                // Trigger notifikasi
                tambahNotif("≡ƒô╕", "Foto dipilih!", input.files[0].name);
            };
            reader.readAsDataURL(input.files[0]);
        }
    }
    </script>
    '''

    return render_page('Tambah Produk + Foto', 'produk', content, extra_scripts)


# =============================================================================
# [PYTHON] ── SSE: REALTIME NOTIFIKASI PESANAN BARU
# =============================================================================
from flask import Response
import time

@app.route('/admin/stream_new_orders')
def stream_new_orders():
    """[PYTHON] SSE Endpoint: Polling database untuk pesanan baru"""
    def generate():
        # Ambil ID pesanan terakhir saat ini
        cur = mysql.connection.cursor()
        cur.execute("SELECT MAX(id) as max_id FROM pesanan")
        res = cur.fetchone()
        last_id = res['max_id'] if res and res['max_id'] else 0
        cur.close()

        while True:
            time.sleep(3)  # Polling setiap 3 detik
            
            cur = mysql.connection.cursor()
            cur.execute("SELECT id, no_pesanan, status FROM pesanan WHERE id > %s ORDER BY id ASC", (last_id,))
            new_orders = cur.fetchall()
            cur.close()

            if new_orders:
                import json
                for order in new_orders:
                    last_id = max(last_id, order['id'])
                    # Kirim data JSON sebagai SSE
                    yield f"data: {json.dumps(order)}\\n\\n"
            else:
                # Kirim komentar agar koneksi tetap hidup (keep-alive)
                yield ": keep-alive\\n\\n"
                
    return Response(generate(), mimetype='text/event-stream')


if __name__ == '__main__':
     

    app.run(
        debug=True,   # debug=True: tampilkan error detail di browser (matikan di production!)
        host='0.0.0.0', # host='0.0.0.0': bisa diakses dari jaringan lokal (bukan hanya localhost)
        port=5000     # port: nomor port server (akses via http://localhost:5000)
    )
    
if __name__ == '__main__':
     

    app.run(
        debug=True,   # debug=True: tampilkan error detail di browser (matikan di production!)
        host='0.0.0.0', # host='0.0.0.0': bisa diakses dari jaringan lokal (bukan hanya localhost)
        port=5000,    # port: nomor port server (akses via http://localhost:5000)
        use_reloader=False # Mencegah error OSError: [WinError 10038] di VS Code Debugger
    )
